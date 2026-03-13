"""
SEC 财报分析器 v6.2 — BeautifulSoup 重构版
修复: PDD 6-K / 中概股 20-F 兼容 + Gross Profit/EPS/BS/Shares 提取
v6.2 修复: AAPL InvCF/CapEx, 股本单位换算, EPS范围检查, G&A模式, 日期回退, DR比例

依赖：pip install beautifulsoup4 lxml openpyxl
"""
import re, sys, csv, argparse
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Tuple
from bs4 import BeautifulSoup
import warnings
try:
    from bs4 import XMLParsedAsHTMLWarning
    warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)
except ImportError:
    pass

DEFAULT_DIR = "sec-data"
OUTPUT_XLSX = "分析结果.xlsx"

@dataclass
class SharesInfo:
    common: Optional[float] = None
    preferred: Optional[float] = None
    weighted_avg: Optional[float] = None
    diluted: Optional[float] = None
    ads_ratio: Optional[str] = None
    ads_per_share: Optional[float] = None  # N ordinary shares per 1 ADS（可为小数，如 VIPS 0.2）
    wtdavg_already_ads: bool = False  # IS 表已以 ADS 为单位报告加权平均股数时为 True（如 IH）
    details: List[Dict] = field(default_factory=list)

@dataclass
class FilingData:
    ticker: str = ""; file_name: str = ""; filing_type: str = ""
    company_name: str = ""; filing_date: str = ""; fiscal_period: str = ""
    currency: str = "USD"; unit_label: str = ""; unit_multiplier: float = 1.0
    revenue: Optional[float] = None; cost_of_revenue: Optional[float] = None
    gross_profit: Optional[float] = None; operating_expenses: Optional[float] = None
    operating_income: Optional[float] = None; net_income: Optional[float] = None
    eps_basic: Optional[float] = None; eps_diluted: Optional[float] = None
    total_assets: Optional[float] = None; current_assets: Optional[float] = None
    cash: Optional[float] = None; total_liabilities: Optional[float] = None
    current_liabilities: Optional[float] = None; long_term_debt: Optional[float] = None
    short_term_debt: Optional[float] = None
    total_equity: Optional[float] = None; retained_earnings: Optional[float] = None
    operating_cf: Optional[float] = None; investing_cf: Optional[float] = None
    financing_cf: Optional[float] = None; capex: Optional[float] = None
    free_cash_flow: Optional[float] = None
    # IS detail
    rd_expense: Optional[float] = None
    sga_expense: Optional[float] = None
    sm_expense: Optional[float] = None
    ga_expense: Optional[float] = None
    interest_income: Optional[float] = None
    interest_expense: Optional[float] = None
    income_tax: Optional[float] = None
    pretax_income: Optional[float] = None
    ebitda: Optional[float] = None
    # BS detail
    short_term_investments: Optional[float] = None
    accounts_receivable: Optional[float] = None
    inventory: Optional[float] = None
    goodwill: Optional[float] = None
    total_non_current_assets: Optional[float] = None
    accounts_payable: Optional[float] = None
    deferred_revenue: Optional[float] = None
    total_non_current_liabilities: Optional[float] = None
    additional_paid_in_capital: Optional[float] = None
    # CF detail
    depreciation_amortization: Optional[float] = None
    stock_based_compensation: Optional[float] = None
    change_in_working_capital: Optional[float] = None
    net_change_in_cash: Optional[float] = None
    cash_end_of_period: Optional[float] = None
    acquisitions: Optional[float] = None
    # Shares & Ratios
    shares: SharesInfo = field(default_factory=SharesInfo)
    gross_margin: Optional[float] = None; operating_margin: Optional[float] = None
    net_margin: Optional[float] = None; current_ratio: Optional[float] = None
    debt_to_equity: Optional[float] = None; roe: Optional[float] = None
    roa: Optional[float] = None
    asset_turnover: Optional[float] = None
    ebitda_margin: Optional[float] = None

# ═══════════ 数字解析 ═══════════
def parse_num(text):
    if not text: return None
    text = text.strip()
    if text in ("-", "\u2014", "\u2013", "N/A", "nil", "None", "$", ""): return None
    neg = "(" in text and ")" in text
    cleaned = re.sub(r"[^0-9.]", "", text)
    if not cleaned or cleaned == ".": return None
    try:
        val = float(cleaned)
        return -val if neg else val
    except ValueError:
        return None

# ═══════════ 文件读取 ═══════════
def read_filing(filepath):
    for enc in ("utf-8", "latin-1", "cp1252"):
        try: return filepath.read_text(encoding=enc)
        except (UnicodeDecodeError, UnicodeError): continue
    return filepath.read_text(encoding="utf-8", errors="replace")

def extract_html_from_sgml(raw):
    if "<DOCUMENT>" not in raw.upper(): return raw
    docs = re.findall(r"<DOCUMENT>(.*?)</DOCUMENT>", raw, re.S | re.I)
    if not docs: return raw
    best, best_n = raw, 0
    for doc in docs:
        n = len(re.findall(r"<table", doc, re.I))
        if n > best_n: best_n, best = n, doc
    others = sorted([d for d in docs if d != best and len(d) > 5000], key=len, reverse=True)
    return best + "\n\n" + "\n\n".join(others[:2])

# ═══════════ 元数据 ═══════════
def detect_filing_type(text, filename):
    fn = filename.upper()
    if re.search(r"10[_-]?K", fn): return "10-K"
    if re.search(r"10[_-]?Q", fn): return "10-Q"
    if re.search(r"20[_-]?F", fn): return "20-F"
    if re.search(r"6[_-]?K", fn): return "6-K"
    t = re.sub(r"<[^>]+>", " ", text[:10000]).upper()
    for k, v in [("FORM 10-K","10-K"),("ANNUAL REPORT","10-K"),("FORM 10-Q","10-Q"),
                 ("FORM 20-F","20-F"),("FORM 6-K","6-K")]:
        if k in t: return v
    return "Unknown"

def extract_company_name_sgml(raw_text):
    m = re.search(r"COMPANY CONFORMED NAME:\s*(.+)", raw_text[:5000])
    if m:
        name = m.group(1).strip()
        if len(name) > 2: return name
    return None

def extract_company_name(text_plain, ticker):
    """公司名：(Exact name of registrant) 上方行 / 全大写行 / ticker"""
    skip = (r"(?i)^(N/?A|FORM|ANNUAL|QUARTER|REPORT|CHECK|INDICATE|TABLE|"
            r"COMMISSION|STATE|IRS|ADDRESS|SECURITIES|DELAWARE|NEVADA|"
            r"CALIFORNIA|MARYLAND|CAYMAN|TRANSITION|PURSUANT|FOR THE|"
            r"BRITISH VIRGIN|IRELAND|HONG KONG|[\d\-]+$)")
    lines = text_plain[:30000].split("\n")
    for i, line in enumerate(lines):
        if re.search(r"exact\s+name\s+of\s+registrant", line, re.I):
            for j in range(i - 1, max(i - 8, -1), -1):
                cand = lines[j].strip()
                if (len(cand) > 2 and not re.match(skip, cand)
                    and not re.search(r"commission|transition|section|pursuant|exchange act|"
                                      r"state or other jurisdiction|incorporation", cand, re.I)):
                    return re.sub(r"\s+", " ", cand)[:80]
            break
    # 全大写行（含 Inc/Ltd/Holdings）
    m = re.search(r"^\s*([A-Z][A-Z0-9\s&.,'-]{4,60}?"
                  r"(?:INC|CORP|LTD|LLC|PLC|GROUP|HOLDINGS|CO|LP|LIMITED)\.?)\s*$",
                  text_plain[:10000], re.M)
    if m: return re.sub(r"\s+", " ", m.group(1).strip())[:80]
    # 混合大小写行（PDD Holdings Inc.）
    m = re.search(r"^\s*([A-Z][A-Za-z0-9\s&.,'-]{4,60}?"
                  r"(?:Inc|Corp|Ltd|LLC|Plc|Group|Holdings|Co|LP|Limited)\.?)\s*$",
                  text_plain[:10000], re.M)
    if m: return re.sub(r"\s+", " ", m.group(1).strip())[:80]
    return ticker or "Unknown"

def detect_currency(text):
    # Only check headers (first 5000 chars) for primary currency
    header = text[:5000].upper()
    for pat, cur in [(r"\bRMB\b|RENMINBI","RMB"),(r"REAIS|\bBRL\b|R\$","BRL"),
                     (r"\bHKD\b|HK\$","HKD"),
                     # EUR: require reporting-currency context; XBRL iso4217:EUR declarations in US filings
                     # should not be counted (many US companies have EUR-denominated subsidiaries/debt).
                     (r"IN\s+(?:THOUSANDS|MILLIONS|BILLIONS)\s+OF\s+EUROS?"
                      r"|(?:EXPRESSED|DENOMINATED|PRESENTED|REPORTED|STATED)\s+IN\s+EUROS?"
                      r"|FUNCTIONAL\s+CURRENCY\s+IS\s+THE\s+EURO", "EUR"),
                     (r"POUND\s*STERLING|\bGBP\b\s+(?:THOUSANDS|MILLIONS|BILLIONS)","GBP"),(r"\bJPY\b","JPY"),
                     # NTD/TWD: require reporting-currency context; bare NTD/TWD in US filings
                     # often refers to Taiwan subsidiaries, not the primary reporting currency.
                     (r"IN\s+(?:THOUSANDS|MILLIONS|BILLIONS)\s+OF\s+(?:NEW\s+TAIWAN\s+)?DOLLARS?\s+\(?NTD\)?"
                      r"|(?:EXPRESSED|DENOMINATED|PRESENTED|REPORTED|STATED)\s+IN\s+(?:NTD\b|TWD\b|NEW\s+TAIWAN\s+DOLLARS?)"
                      r"|FUNCTIONAL\s+CURRENCY\s+IS\s+THE\s+(?:NTD|TWD|NEW\s+TAIWAN\s+DOLLAR)"
                      r"|iso4217:TWD", "NTD"),
                     (r"\bINR\b|INDIAN\s+RUPEE|&#8377;|₹","INR")]:
        if re.search(pat, header): return cur
    # Broader check with strict patterns (avoid "yen" in body text)
    t = text[:50000].upper()
    # 若明确声明以美元报告，提前返回 USD（防止后续扩展搜索误匹配）
    if re.search(r"(?:U\.S\.|UNITED\s+STATES)\s+DOLLARS?\s+IN\s+(?:THOUSANDS|MILLIONS|BILLIONS)"
                 r"|IN\s+(?:THOUSANDS|MILLIONS|BILLIONS)\s+OF\s+(?:U\.S\.|UNITED\s+STATES)\s+DOLLARS?"
                 r"|(?:EXPRESSED|DENOMINATED|PRESENTED|REPORTED|STATED)\s+IN\s+(?:U\.S\.|UNITED\s+STATES)\s+DOLLARS?"
                 r"|FUNCTIONAL\s+CURRENCY\s+IS\s+THE\s+(?:U\.S\.|UNITED\s+STATES)\s+DOLLAR", t):
        return "USD"
    for pat, cur in [(r"\bRMB\b|RENMINBI","RMB"),(r"REAIS|\bBRL\b","BRL"),
                     (r"\bHKD\b|HK\$","HKD"),
                     # GBP：要求在报告货币上下文中出现，排除括号折算值如 "(GBP 1.8 million)"
                     (r"POUND\s*STERLING"
                      r"|IN\s+(?:THOUSANDS|MILLIONS|BILLIONS)\s+OF\s+(?:BRITISH\s+)?POUNDS?"
                      r"|(?:EXPRESSED|DENOMINATED|PRESENTED|REPORTED|STATED)\s+IN\s+(?:GBP\b|BRITISH\s+POUNDS?|POUNDS?\s+STERLING)"
                      r"|FUNCTIONAL\s+CURRENCY\s+IS\s+THE\s+(?:GBP|BRITISH\s+POUND|POUND\s+STERLING)", "GBP"),
                     (r"DENOMINATED\s+IN.*?(?:YEN|JPY)|\bJPY\b","JPY"),
                     (r"IN\s+(?:THOUSANDS|MILLIONS|BILLIONS)\s+OF\s+(?:NEW\s+TAIWAN\s+)?DOLLARS?\s+\(?NTD\)?"
                      r"|(?:EXPRESSED|DENOMINATED|PRESENTED|REPORTED|STATED)\s+IN\s+(?:NTD\b|TWD\b|NEW\s+TAIWAN\s+DOLLARS?)"
                      r"|FUNCTIONAL\s+CURRENCY\s+IS\s+THE\s+(?:NTD|TWD|NEW\s+TAIWAN\s+DOLLAR)"
                      r"|iso4217:TWD", "NTD"),
                     # INR: Indian Rupee — match explicit declarations, exclude FX rate mentions like "USD/INR at 84"
                     (r"INDIAN\s+RUPEE|&#8377;|₹"
                      r"|IN\s+(?:THOUSANDS|MILLIONS|BILLIONS)\s+OF\s+(?:INDIAN\s+)?RUPEES?"
                      r"|(?:EXPRESSED|DENOMINATED|PRESENTED|REPORTED|STATED)\s+IN\s+(?:INR\b|INDIAN\s+RUPEES?)"
                      r"|FUNCTIONAL\s+CURRENCY\s+IS\s+THE\s+(?:INR|INDIAN\s+RUPEE)"
                      r"|iso4217:INR", "INR")]:
        if re.search(pat, t): return cur
    # Extended check (up to 1MB) for filings where currency is declared deep in the document
    t2 = text[:1000000].upper()
    for pat, cur in [
        # CAD: 仅匹配明确的报告货币声明，排除外汇风险列表中的提及
        (r"iso4217:CAD"
         r"|(?:EXPRESSED|DENOMINATED|REPORTED|STATED|PRESENTED)\s+IN\s+(?:CANADIAN\s+DOLLARS?|CAD\b)"
         r"|(?:IN\s+MILLIONS|IN\s+THOUSANDS)\s+OF\s+CANADIAN\s+DOLLARS?"
         r"|AMOUNTS?\s+(?:ARE\s+)?IN\s+(?:CANADIAN\s+DOLLARS?|CAD\b)"
         r"|FUNCTIONAL\s+CURRENCY\s+IS\s+THE\s+CANADIAN\s+DOLLAR", "CAD"),
        (r"IN\s+(?:THOUSANDS|MILLIONS|BILLIONS)\s+OF\s+(?:NEW\s+TAIWAN\s+)?DOLLARS?\s+\(?NTD\)?"
         r"|(?:EXPRESSED|DENOMINATED|PRESENTED|REPORTED|STATED)\s+IN\s+(?:NTD\b|TWD\b|NEW\s+TAIWAN\s+DOLLARS?)"
         r"|FUNCTIONAL\s+CURRENCY\s+IS\s+THE\s+(?:NTD|TWD|NEW\s+TAIWAN\s+DOLLAR)"
         r"|iso4217:TWD", "NTD"),
        # RMB: in 1MB scan require reporting-currency context to avoid false positives
        # from FX-risk disclosures listing "Chinese Renminbi" alongside many other currencies.
        (r"IN\s+(?:THOUSANDS|MILLIONS|BILLIONS)\s+OF\s+(?:CHINESE\s+)?(?:YUAN|RMB|RENMINBI)"
         r"|(?:EXPRESSED|DENOMINATED|PRESENTED|REPORTED|STATED)\s+IN\s+(?:RMB\b|RENMINBI|YUAN)"
         r"|FUNCTIONAL\s+CURRENCY\s+IS\s+(?:THE\s+)?(?:RMB|RENMINBI|CHINESE\s+YUAN)"
         r"|iso4217:CNY", "RMB"),
        (r"\bHKD\b|HK\$","HKD")]:
        if re.search(pat, t2): return cur
    return "USD"

def extract_date_info(text):
    fd = fp = ""
    for p in [r"(?:filed|date\s*of\s*report|filing\s*date)[:\s]*(\w+\s+\d{1,2},?\s*\d{4})",
              r"(?:dated?|as\s+of)[:\s]*(\w+\s+\d{1,2},?\s*\d{4})"]:
        m = re.search(p, text[:20000], re.I)
        if m: fd = m.group(1).strip(); break
    for p in [r"(?:fiscal\s*year|year|period|twelve\s*months?|(?:three|six|nine)\s*months?)\s*ended?\s*:?\s*(\w+\s+\d{1,2},?\s*\d{4})",
              r"for\s+the\s+(?:fiscal\s+)?(?:year|quarter)\s+ended\s+(\w+\s+\d{1,2},?\s*\d{4})",
              r"(?:quarter|period)\s+ended\s+(\w+\s+\d{1,2},?\s*\d{4})"]:
        m = re.search(p, text[:20000], re.I)
        if m: fp = m.group(1).strip(); break
    return fd, fp

def date_from_filename(filename):
    """从文件名 TICKER_FORM_YYYY-MM-DD.ext 提取日期，转换为 'Month D, YYYY' 格式"""
    m = re.search(r'(\d{4})-(\d{2})-(\d{2})', str(filename))
    if not m: return ""
    try:
        from datetime import datetime
        dt = datetime.strptime(m.group(0), "%Y-%m-%d")
        return dt.strftime("%B %-d, %Y")   # macOS/Linux; Windows 用 %#d
    except Exception:
        return ""

# ═══════════ HTML 表格引擎 ═══════════
def parse_table_to_rows(table):
    rows = []
    for tr in table.find_all("tr"):
        cells = []
        for td in tr.find_all(["td", "th"]):
            text = td.get_text(strip=True)
            colspan = int(td.get("colspan", 1))
            cells.append(text)
            cells.extend([""] * (colspan - 1))
        rows.append(cells)
    return rows

def detect_table_unit(table, raw_html):
    ht = ""
    for tr in table.find_all("tr")[:8]:
        ht += " " + tr.get_text(separator=" ", strip=True)
    hu = ht.upper()
    if re.search(r"IN\s+THOUSANDS", hu): return "thousands", 0.001
    if re.search(r"IN\s+(?:\w+\s+)?MILLIONS?(?:\b|$|[^A-Z])", hu): return "millions", 1.0
    if re.search(r"IN\s+(?:\w+\s+)?BILLIONS?(?:\b|$|[^A-Z])", hu): return "billions", 1000.0
    # "IN USD" / "IN U.S. DOLLARS" / "IN US DOLLARS" without scale qualifier = individual dollars
    if re.search(r"\bIN\s+(?:USD|U\.S\.\s+DOLLARS?|US\s+DOLLARS?)\b", hu):
        return "ones", 1e-6
    # "$'000" / "US$'000" / "HK$'000" notation = thousands (common in HK/AU/SG filings)
    # Note: the apostrophe may be a Unicode curly quote (U+2019 ') not straight apostrophe
    if re.search(r"(?:US\$|HK\$|AU\$|NZ\$|\$|USD|RMB|CNY|SGD)\s*[\u2019\u2018'\u0060]?\s*000\b", ht):
        return "thousands(notation)", 0.001
    # Method 2b: XBRL inline decimals/scale attributes — checked BEFORE nearby-HTML scan.
    # ix:nonFraction decimals="-3" → thousands; "-6" → millions; "0" → ones.
    # When scale attribute is present (iXBRL): display_unit = 10^scale (scale=6 → millions,
    # scale=3 → thousands, etc.). Scale overrides decimals for determining displayed unit.
    try:
        dec_vals = []
        scale_vals = []
        import math
        for elem in table.find_all(True, attrs={"decimals": True}):
            try:
                v = float(elem["decimals"])
                if not math.isinf(v): dec_vals.append(int(v))
            except: pass
            try:
                sc = elem.get("scale")
                if sc is not None:
                    sv = int(sc)
                    scale_vals.append(sv)
            except: pass
        if dec_vals or scale_vals:
            from collections import Counter
            if scale_vals:
                # scale attribute determines the display unit: displayed = raw × 10^scale
                most_common_scale = Counter(scale_vals).most_common(1)[0][0]
                if most_common_scale == 6: return "millions(xbrl)", 1.0
                if most_common_scale == 3: return "thousands(xbrl)", 0.001
                if most_common_scale == 0: return "ones(xbrl)", 1e-6
                if most_common_scale == 9: return "billions(xbrl)", 1000.0
            if dec_vals:
                most_common = Counter(dec_vals).most_common(1)[0][0]
                if most_common == -3: return "thousands(xbrl)", 0.001
                if most_common == -6: return "millions(xbrl)", 1.0
                if most_common == 0:  return "ones(xbrl)", 1e-6
                if most_common == -9: return "billions(xbrl)", 1000.0
    except Exception: pass
    # Method 4 (early): Currency-symbol column headers + large integer values → ones
    # Run BEFORE nearby/DOM scan to prevent false "thousands" from DOM overriding clear "ones" evidence.
    # (e.g., GMHS: "$" data cells with values like 15,328,252 → individual dollars)
    # Guard: only fires when standalone $ is present AND numeric values look like individual dollars
    # (i.e., at least one value >= 1,000,000 — distinguishes $15M individual from $8,492.6 millions)
    try:
        _hdr_cells_early = []
        for tr in table.find_all("tr")[:6]:
            for td in tr.find_all(["td", "th"]):
                cell = td.get_text(" ", strip=True).strip().upper()
                if cell: _hdr_cells_early.append(cell)
        _hdr_txt_early = " ".join(_hdr_cells_early)
        if not re.search(r"THOUSANDS|MILLIONS|BILLIONS", _hdr_txt_early):
            _has_dollar = (any(c in ("US$", "$", "HK$", "C$") for c in _hdr_cells_early) or
                           any(t in ("US$", "$", "HK$", "C$")
                               for cell in _hdr_cells_early for t in cell.split()))
            if _has_dollar:
                # Check if any numeric value in first 10 rows is >= 1,000,000 (individual dollars)
                _has_large_val = False
                for _tr in table.find_all("tr")[:10]:
                    _row_txt = _tr.get_text(" ", strip=True)
                    for _m in re.finditer(r'\b(\d[\d,]+)\b', _row_txt):
                        try:
                            _v = float(_m.group(1).replace(",", ""))
                            if _v >= 1_000_000:
                                _has_large_val = True; break
                        except: pass
                    if _has_large_val: break
                if _has_large_val:
                    return "ones", 1e-6
    except Exception: pass
    # Method 2: Search raw_html before the table for unit qualifiers.
    # Anchor strategy: use the first single text node that looks like a financial label
    # (not a date/year cell), normalized for non-breaking spaces and case.
    try:
        tbl_texts = [t.strip().replace('\xa0', ' ').replace('\u2009', ' ')
                     for t in table.stripped_strings if t.strip()]
        # Pick first text that is a financial label (not a date string or purely numeric)
        # Exclude generic temporal headers that appear in many tables (e.g. "Year Ended December 31,")
        # to avoid matching the wrong table in the raw HTML when the same header repeats.
        anchor_text = ''
        for tok in tbl_texts:
            if (len(tok) >= 8 and re.search(r'[A-Za-z]{3}', tok)
                    and not re.match(r'^[\d\s,./]+$', tok)
                    and not re.match(r'^(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)', tok, re.I)
                    and not re.match(r'^(?:first|second|third|fourth|q[1-4])\s', tok, re.I)
                    and not re.match(r'^year\s+ended', tok, re.I)
                    and not re.match(r'^(?:three|six|nine|twelve)\s+months', tok, re.I)
                    and not re.match(r'^for\s+the\s+(?:three|six|nine|twelve|year)', tok, re.I)
                    and not re.match(r'^as\s+of\b', tok, re.I)):
                anchor_text = tok[:50]
                break
        pos = -1
        if anchor_text:
            pos = raw_html.find(anchor_text)
            if pos < 0:
                pos = raw_html.upper().find(anchor_text.upper())
        if pos < 0:
            # Fallback: str(table) anchor (works for lxml-normalized htm files)
            anchor = str(table)[:80]
            pos = raw_html.find(anchor[:60])
        if pos > 0:
            before = re.sub(r"<[^>]+>", " ", raw_html[max(0, pos-20000):pos]).upper()
            tp = mp = -1
            for m in re.finditer(r"IN\s+THOUSANDS", before): tp = m.start()
            for m in re.finditer(r"IN\s+MILLIONS", before): mp = m.start()
            if tp > mp >= 0 or (tp >= 0 and mp < 0): return "thousands(nearby)", 0.001
            if mp > tp >= 0 or (mp >= 0 and tp < 0): return "millions(nearby)", 1.0
            # Also search after the anchor (some filings put unit note after the table header)
            after = re.sub(r"<[^>]+>", " ", raw_html[pos:min(len(raw_html), pos+5000)]).upper()
            tp2 = mp2 = -1
            for m in re.finditer(r"IN\s+THOUSANDS", after): tp2 = m.start()
            for m in re.finditer(r"IN\s+MILLIONS", after): mp2 = m.start()
            if tp2 >= 0 and (mp2 < 0 or tp2 < mp2): return "thousands(nearby-after)", 0.001
            if mp2 >= 0 and (tp2 < 0 or mp2 < tp2): return "millions(nearby-after)", 1.0
    except Exception: pass
    # Method 3: Check preceding sibling/parent elements in DOM
    # Only use if the found text is a close ancestor/sibling of the table (within 4 levels)
    try:
        prev = table.find_previous(string=re.compile(r"[Ii]n\s+(?:millions|thousands|billions)"))
        if prev:
            # Collect close ancestors of table (up to 4 levels)
            tbl_close = set()
            p = table.parent
            for _ in range(4):
                if p: tbl_close.add(id(p)); p = p.parent
            # prev is "in scope" if any of its ancestors (up to 4 levels) is a close ancestor of table
            prev_in_scope = any(id(a) in tbl_close for a in list(prev.parents)[:4])
            if prev_in_scope:
                ptxt = prev.strip().upper()
                if "IN MILLIONS" in ptxt: return "millions(dom)", 1.0
                if "IN THOUSANDS" in ptxt: return "thousands(dom)", 0.001
                if "IN BILLIONS" in ptxt: return "billions(dom)", 1000.0
    except Exception: pass
    # Method 4: Currency-symbol column headers without unit qualifier
    # (e.g., YALA: "US$" standalone cells; NFGC: "$" embedded in "Sep 30, 2025 $" cells)
    try:
        header_cells = []
        for tr in table.find_all("tr")[:6]:
            for td in tr.find_all(["td", "th"]):
                cell = td.get_text(" ", strip=True).strip().upper()
                if cell: header_cells.append(cell)
        hdr_txt = " ".join(header_cells)
        if not re.search(r"THOUSANDS|MILLIONS|BILLIONS", hdr_txt):
            _has_currency = (
                any(c in ("US$", "$", "HK$", "C$") for c in header_cells) or
                any(t in ("US$", "$", "HK$", "C$")
                    for cell in header_cells for t in cell.split())
            )
            if _has_currency:
                # Magnitude check: ones tables have raw values >= 1,000,000 (e.g. $950M = 950,000,000).
                # Millions tables show the same amount as "950.0", max < 1,000,000.
                # Without this check, tables reporting in $M with standalone "$" cells (AMG, APD, JHX)
                # are incorrectly classified as ones, giving results 1,000,000x too small.
                _has_large_val_m4 = False
                for _tr_m4 in table.find_all("tr")[:10]:
                    for _m_m4 in re.finditer(r'\b(\d[\d,]+)\b', _tr_m4.get_text(" ", strip=True)):
                        try:
                            if float(_m_m4.group(1).replace(",", "")) >= 1_000_000:
                                _has_large_val_m4 = True; break
                        except: pass
                    if _has_large_val_m4: break
                if _has_large_val_m4:
                    return "ones", 1e-6
    except Exception: pass
    # Method 2c: Full-document fallback — scan entire doc for unit declarations
    # Last resort when table header and nearby HTML have no explicit unit marker.
    # Only fires when all other methods have returned nothing.
    try:
        full_text = re.sub(r"<[^>]+>", " ", raw_html).upper()
        tp = mp = -1
        for m in re.finditer(r"IN\s+THOUSANDS", full_text): tp = m.start()
        for m in re.finditer(r"IN\s+MILLIONS", full_text): mp = m.start()
        if tp > mp >= 0 or (tp >= 0 and mp < 0): return "thousands(doc)", 0.001
        if mp > tp >= 0 or (mp >= 0 and tp < 0): return "millions(doc)", 1.0
    except Exception: pass
    return "unknown", 0

def classify_table(table):
    text = table.get_text(separator=" ", strip=True).lower()
    if len(text) < 150: return None
    is_s = bs_s = cf_s = 0
    # IS 指标
    if re.search(r"(?:total )?(?:net )?revenues?(?:\s|$)", text): is_s += 2
    if re.search(r"gross profit", text): is_s += 3
    if re.search(r"(?:total )?(?:operating )?expenses", text): is_s += 1
    if re.search(r"(?:income|loss) from operations|operating (?:income|loss|profit)", text): is_s += 2
    if re.search(r"cost of (?:revenues?|goods|sales|services)", text): is_s += 2
    if re.search(r"(?:basic|diluted).*?per (?:share|ads|common|ordinary)", text): is_s += 3
    if re.search(r"net (?:income|loss|earnings|profit)", text): is_s += 1
    if re.search(r"income\s+tax|provision\s+for\s+income", text): is_s += 1
    # BS 指标
    if re.search(r"total assets", text): bs_s += 3
    if re.search(r"total liabilities", text): bs_s += 3
    if re.search(r"(?:stockholders|shareholders|owners).*?equity", text): bs_s += 3
    if re.search(r"(?:total )?current assets", text): bs_s += 2
    if re.search(r"(?:total )?current liabilities", text): bs_s += 2
    if re.search(r"accounts?\s+(?:receivable|payable)", text): bs_s += 1
    if re.search(r"goodwill|intangible", text): bs_s += 1
    # CF 指标
    if re.search(r"operating activities", text): cf_s += 4
    if re.search(r"investing activities", text): cf_s += 4
    if re.search(r"financing activities", text): cf_s += 4
    if re.search(r"depreciation", text): cf_s += 2
    if re.search(r"cash.*?(?:beginning|end) of (?:period|year)", text): cf_s += 3
    if re.search(r"stock.based compensation", text): cf_s += 1
    if re.search(r"share.based compensation", text): cf_s += 1
    # CF 高分压制 IS
    if cf_s >= 8: is_s = max(0, is_s - 5)
    scores = {"IS": is_s, "BS": bs_s, "CF": cf_s}
    best = max(scores, key=scores.get)
    if scores[best] < 4: return None
    if is_s >= 4 and bs_s >= 4:
        return "IS" if re.search(r"per share|per ads|per ordinary", text) else ("BS" if bs_s > is_s else "IS")
    return best

# ═══════════ 行级提取 ═══════════
def extract_numbers_from_row(cells):
    raw_nums = []
    i = 0
    cells_list = list(cells)
    while i < len(cells_list):
        c = cells_list[i].strip()
        if not c or c in ("$", "RMB", "US$", "R$"):
            i += 1; continue
        # Strip trailing footnote references to avoid treating note numbers as data
        c = re.sub(r"\s*\(notes?\s+[\d,\s\w]+\)\s*$", "", c, flags=re.I).strip()
        if not c: i += 1; continue
        # Handle split parentheses: "(2.17" + ")" across adjacent cells
        if c.startswith("(") and ")" not in c and i + 1 < len(cells_list) and cells_list[i + 1].strip() == ")":
            c = c + ")"
            i += 1  # consume the ")" cell too
        v = parse_num(c)
        if v is not None: raw_nums.append(v)
        i += 1
    if not raw_nums: return raw_nums
    # Filter footnote refs: small first number + much larger subsequent numbers
    if len(raw_nums) >= 2 and 0 < abs(raw_nums[0]) < 50:
        max_rest = max(abs(v) for v in raw_nums[1:]) if raw_nums[1:] else 0
        # Case 1: footnote tiny vs large financial numbers
        if max_rest > abs(raw_nums[0]) * 100:
            return raw_nums[1:]
        # Case 2: footnote is integer, rest are small decimals (EPS)
        if raw_nums[0] == int(raw_nums[0]) and all(abs(v) < 100 for v in raw_nums[1:]):
            if any(v != int(v) for v in raw_nums[1:] if v != 0):
                return raw_nums[1:]
    return raw_nums

def build_row_label(cells):
    parts = []
    for c in cells[:6]:
        c = c.strip()
        if not c or c in ("$", "RMB", "US$", "R$"): continue
        # Strip trailing footnote references before numeric check to avoid false positives
        c_clean = re.sub(r"\s*\(notes?\s+[\d,\s\w]+\)\s*$", "", c, flags=re.I).strip()
        v = parse_num(c_clean)
        if v is not None and abs(v) > 50: break
        parts.append(c_clean)
    label = " ".join(parts)
    # Strip trailing small integers (legacy fallback)
    label = re.sub(r"\s+\d{1,2}$", "", label)
    return label

def find_row_value(rows, patterns, col_index=0, negate_loss=False):
    for pat in patterns:
        for row in rows:
            if not row: continue
            label = build_row_label(row)
            try:
                if re.search(pat, label, re.I):
                    nums = extract_numbers_from_row(row)
                    if nums and col_index < len(nums):
                        val = nums[col_index]
                        if negate_loss and val is not None and val > 0:
                            # Pure "loss" label (no "income"/"profit"/"earnings" alongside)
                            # e.g. "Operating loss", "Net loss", "Loss from operations"
                            ll = label.lower()
                            if re.search(r'\bloss\b', ll) and not re.search(r'\b(?:income|profit|earning)\b', ll):
                                val = -val
                        return val
            except re.error: continue
    return None

# Fields whose matched-label "loss" rows should yield negative values
_NEGATE_LOSS_FIELDS = {"operating_income", "net_income", "pretax_income", "gross_profit"}

def extract_from_table(rows, patterns, col=0):
    return {k: find_row_value(rows, pats, col_index=col,
                              negate_loss=(k in _NEGATE_LOSS_FIELDS))
            for k, pats in patterns.items()}

def detect_latest_column(rows):
    """
    检测表格中最新数据所在的列索引。
    使用年份+月份作为排序键，找到最新日期对应的第一个数据列。
    例：'2023 | 2024 | 2025' → col 2；'2025 | 2024 | 2023' → col 0；
    'Sep-2024 | Jun-2025 | Sep-2025' → col 2（月份感知，避免选到 Jun-2025）。
    跨行关联：年份和月份可分布在不同的行（如 NTES：第1行含月份，第2行含年份）。
    """
    _MONTH_MAP = {
        'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
        'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12,
        'january':1,'february':2,'march':3,'april':4,'june':6,'july':7,
        'august':8,'september':9,'october':10,'november':11,'december':12,
    }
    header_rows = rows[:6]
    # Pass 1: build per-table-column-index maps for year and month (cross-row)
    # table_col_year[k] = first year found in column k across all header rows
    # table_col_month[k] = month found in column k across all header rows
    table_col_year = {}
    table_col_month = {}
    for row in header_rows:
        for k, cell in enumerate(row):
            ct = cell.strip()
            if not ct:
                continue
            years = re.findall(r'20[12][0-9]', ct)
            if years and k not in table_col_year:
                table_col_year[k] = int(years[0])
            cl = ct.lower()
            if k not in table_col_month:
                for mname, mnum in _MONTH_MAP.items():
                    if mname in cl:
                        table_col_month[k] = mnum
                        break
    # Pass 2: build date_positions list using cross-row month lookup when month=0
    date_positions = []  # [(sortkey, col_pos)]
    for row in header_rows:
        col_pos = 0
        for k, cell in enumerate(row):
            ct = cell.strip()
            if not ct:
                continue
            years_in_cell = re.findall(r'20[12][0-9]', ct)
            for y in years_in_cell:
                # First try month from this cell; fall back to same column in other rows
                month = 0
                cl = ct.lower()
                for mname, mnum in _MONTH_MAP.items():
                    if mname in cl:
                        month = mnum
                        break
                if month == 0:
                    month = table_col_month.get(k, 0)
                date_positions.append((int(y) * 100 + month, col_pos))
            v = parse_num(ct)
            if v is not None and abs(v) > 50:
                col_pos += 1

    if not date_positions:
        return 0

    max_key = max(dp[0] for dp in date_positions)
    if max_key == min(dp[0] for dp in date_positions):
        return 0  # only one date found, use first column

    # Return col_pos of the FIRST occurrence of the max date key
    for key, cp in date_positions:
        if key == max_key:
            return cp
    return 0

def extract_eps_contextual(rows, col=0):
    """Context-aware EPS extraction for filings with '-Basic'/'-Diluted' sub-rows."""
    eps_b = eps_d = None
    in_eps_section = False
    in_shares_section = False
    for row in rows:
        # Get first non-empty cell as raw label
        first_cell = ''
        for c in row:
            c = c.strip()
            if c and c not in ('', '$', 'RMB', 'US$'):
                first_cell = c
                break
        fl = first_cell.lower().strip()
        # Detect section headers
        if re.search(r'(?:earnings|income|loss).*per\s+(?:share|ads|ordinary|common)', fl):
            in_eps_section = True
            in_shares_section = False
            continue
        if re.search(r'weighted', fl):
            in_eps_section = False
            in_shares_section = True
            continue
        # Match -Basic / -Diluted (also handles em-dash — and en-dash – and suffix ——basic)
        is_basic = re.match(r'^[-—–]?\s*basic', fl) or re.search(r'[—–-]{1,2}\s*basic\s*$', fl)
        is_diluted = re.match(r'^[-—–]?\s*diluted', fl) or re.search(r'[—–-]{1,2}\s*diluted\s*$', fl)
        if is_basic:
            nums = extract_numbers_from_row(row)
            _v = nums[col] if (nums and len(nums) > col) else (nums[0] if nums else None)
            if _v is not None and in_eps_section and abs(_v) < 1000 and eps_b is None:
                eps_b = _v
        if is_diluted:
            nums = extract_numbers_from_row(row)
            _v = nums[col] if (nums and len(nums) > col) else (nums[0] if nums else None)
            if _v is not None and in_eps_section and abs(_v) < 1000 and eps_d is None:
                eps_d = _v
    return eps_b, eps_d

def extract_shares_contextual(rows, col=0, table_text=""):
    """Context-aware shares extraction for '-Basic'/'-Diluted' sub-rows under Weighted-average header.
    Returns (sh_basic, sh_diluted, in_thousands, already_ads)
    already_ads=True when the weighted avg label explicitly says 'ADSs' (e.g. iHuman)."""
    sh_b = sh_d = None
    in_shares = False
    already_ads = False
    _tl = table_text.lower()
    in_thousands = "thousand" in _tl
    # 若表头注明 "except (for) share/per-share data"，股数是原始计数而非 thousands
    if in_thousands and re.search(r"except\s+(?:for\s+)?(?:share|per[\s\-]share)", _tl):
        in_thousands = False
    for row in rows:
        first_cell = ''
        for c in row:
            c = c.strip()
            if c and c not in ('', '$', 'RMB', 'US$'):
                first_cell = c; break
        fl = first_cell.lower()
        if ('weighted' in fl or 'shares used' in fl) and ('share' in fl or 'ordinary' in fl or 'ads' in fl or 'computing' in fl):
            in_shares = True
            # 若标签明确写 "ADSs" / "ADS"，说明 IS 已用 ADS 单位报告加权平均股数
            if re.search(r'\bads[s]?\b', fl):
                already_ads = True
            continue
        if in_shares and fl and not re.match(r'^[-—–]{0,2}\s*(?:basic|diluted)', fl):
            # blank rows (fl='') are spacer rows — do NOT reset in_shares
            in_shares = False
            continue
        if in_shares and re.match(r'^[-—–]{0,2}\s*basic\s+and\s+diluted', fl):
            # Loss company: combined "Basic and diluted" row
            nums = extract_numbers_from_row(row)
            _v = nums[col] if (nums and len(nums) > col) else (nums[0] if nums else None)
            if _v is not None and _v > 1000:
                if sh_b is None: sh_b = _v
                if sh_d is None: sh_d = _v
            continue
        if in_shares and re.match(r'^[-—–]{0,2}\s*basic', fl):
            nums = extract_numbers_from_row(row)
            _v = nums[col] if (nums and len(nums) > col) else (nums[0] if nums else None)
            if _v is not None and _v > 1000 and sh_b is None:
                sh_b = _v
        if in_shares and re.match(r'^[-—–]{0,2}\s*diluted', fl):
            nums = extract_numbers_from_row(row)
            _v = nums[col] if (nums and len(nums) > col) else (nums[0] if nums else None)
            if _v is not None and _v > 1000 and sh_d is None:
                sh_d = _v
    return sh_b, sh_d, in_thousands, already_ads

# ═══════════ 标签模式 ═══════════
IS_PATTERNS = {
    "revenue": [r"^total\s+revenue(?:\s+and\s+income)?$", r"^total\s+(?:net\s+)?revenues?$",
                r"^(?:total\s+)?(?:net\s+)?revenues?$", r"^(?:total\s+)?net\s+sales$",
                r"^revenues?$",
                r"^total\s+operating\s+revenues?$",         # LX, ATO
                r"^total\s+net\s+revenue$",                # QFIN
                r"^total\s+revenue\s+and\s+income\b",      # STNE
                r"^total\s+revenues?\s*\(excluding\b",     # JFU
                r"^sales$",                                # MRK: bare "Sales" label
                r"^total\s+sales$",
],
    "cost_of_revenue": [r"^(?:total\s+)?costs?\s+of\s+(?:revenues?|goods\s+sold|sales|services)",
                        r"^total\s+cost\s+of\s+revenue",
                        r"^cost\s+of\s+services"],
    "gross_profit": [r"^gross\s+profit$", r"^total\s+gross\s+profit$",
                     r"^gross\s+(?:profit|margin)\b",
                     r"^net\s+interest\s+income$",                   # RY/banks: Net interest income
],
    "operating_expenses": [r"^total\s+(?:costs?\s+and\s+)?operating\s+expenses$",
                           r"^total\s+operating\s+(?:costs?\s+and\s+)?expenses$",
                           r"^total\s+(?:costs?\s+and\s+)?expenses$",
                           r"^non.interest\s+expense$",              # RY/banks
],
    "operating_income": [r"^(?:income|loss|profit)\s*(?:\(loss\))?\s*from\s+operations$",
                         r"^operating\s+(?:income|loss|profit)",
                         r"^income\s+\(loss\)\s+from\s+operations$",
                         r"^(?:total\s+)?operating\s+(?:income|profit|loss)",
                         r"^[\(/]?\s*(?:loss|income|profit)[\)/]?(?:[/\s]+(?:income|loss|profit))?\s*from\s+operations",  # TAL
                         r"^pre.provision,?\s*pre.tax\s*earnings",    # RY/banks: Pre-provision, pre-tax earnings
],
    "net_income": [r"^net\s+(?:income|loss|earnings?|profit)(?:\s+and\s+comprehensive)?(?:\s+\(?(?:income|loss)\)?)?\s*$",
                   r"^net\s+(?:income|loss)\s+attributable\s+to",
                   r"^(?:profit|loss)\s+for\s+the\s+(?:\w+\s+)?(?:year|period|quarter)",
                   r"^net\s+(?:income|loss|earnings?|profit)",
                   r"^net\s+(?:income|loss)\s+for\s+the\s+(?:period|quarter|year)",
                   r"^(?:profit|loss)\s+attributable\s+to",
                   r"^net\s+\(?loss\)?\s+(?:income|profit)",  # GOOS/JFU: "Net (loss) income"
],
    "eps_basic": [r"(?:basic|—\s*basic)\s*(?:net\s+)?(?:income|earnings?|loss|profit)\s*per\s*(?:share|ads|ordinary|common)",
                  r"^basic\s+earnings?\s+per\s+(?:common\s+)?share",
                  r"(?:net\s+)?(?:income|earnings?)\s+per\s+(?:share|ads|ordinary).*?basic",
                  r"basic\s+(?:income|earnings?)\s+per\s+(?:common\s+)?(?:share|ads|ordinary)",
                  r"per\s+(?:ordinary\s+)?share.*?basic",
                  r"per\s+ADS.*?basic",
],
    "eps_diluted": [r"(?:diluted|—\s*diluted)\s*(?:net\s+)?(?:income|earnings?|loss|profit)\s*per\s*(?:share|ads|ordinary|common)",
                    r"^diluted\s+earnings?\s+per\s+(?:common\s+)?share",
                    r"(?:net\s+)?(?:income|earnings?)\s+per\s+(?:share|ads|ordinary).*?diluted",
                    r"diluted\s+(?:income|earnings?)\s+per\s+(?:common\s+)?(?:share|ads|ordinary)",
                    r"per\s+(?:ordinary\s+)?share.*?diluted",
                    r"per\s+ADS.*?diluted",
],
    "rd_expense": [r"^research\s+and\s+development"],
    "sm_expense": [r"^sales\s+and\s+marketing", r"^selling\s+and\s+marketing",
                   r"^selling\s+expenses?"],
    "ga_expense": [r"^general\s+and\s+administrative\b(?!.*selling)",   # 纯 G&A，不含 Selling 前缀
                   r"^administrative\s+expenses?"],
    "sga_expense": [r"^selling,?\s+general\s+and\s+administrative"],    # Apple 等公司：Selling, G&A 合并行
    "interest_income": [r"^interest\s+(?:income|and\s+investment\s+income)", r"^interest\s+income"],
    "interest_expense": [r"^interest\s+expense"],
    "income_tax": [r"^(?:\(?benefit\s+from\)?\s+)?(?:provision\s+for\s+)?income\s+tax",
                    r"^(?:provision\s+for|income\s+tax)", r"^income\s+tax\s+expense",
                    r"^income\s+tax\s+and\s+social\s+contribution"],
    "pretax_income": [r"^(?:income|loss|profit)\s*(?:\(loss\)\s*)?\s*before\s+(?:income\s+)?tax",
                       r"^(?:income|profit)\s+before\s+(?:provision|income\s+tax)",
                       r"^profit\s+before\s+income\s+tax",
                       r"^[\(/]?\s*(?:loss|income|profit)[\)/]?\s*(?:income|loss|profit)?\s*before\s+(?:income\s+)?tax",  # JFU: (Loss) income before income tax
],
}
SHARES_PATTERNS = {
    "shares_basic": [r"weighted[\s-]*average.*?(?:basic|computing|outstanding|shares)",
                     r"denominator.*?basic",
                     r"(?:basic|ordinary)\s+(?:weighted|shares)",
                     r"shares\s+used\s+to\s+compute\s+basic",
                     r"weighted\s+average\s+number\s+of\s+(?:outstanding\s+)?(?:common\s+)?shares",
                     # "Basic average shares outstanding" format (e.g. NUE) — no "weighted" keyword
                     r"basic\s+average\s+shares",
                     r"average\s+(?:common\s+)?shares?\s+outstanding.*?basic"],
    "shares_diluted": [r"denominator.*?dilut(?:ive|ed).*?(?:weighted|share)",
                       r"weighted[\s-]*average.*?dilut",
                       r"diluted\s+(?:weighted|shares)",
                       r"shares\s+used\s+to\s+compute\s+diluted",
                       # "Diluted average shares outstanding" format (e.g. NUE)
                       r"diluted\s+average\s+shares",
                       r"average\s+(?:common\s+)?shares?\s+outstanding.*?diluted"],
}
BS_PATTERNS = {
    "total_assets": [r"^total\s+assets$", r"^total$",
                     r"^total\s+assets\s+and\s+liabilities$",
                     r"^total\s+liabilities\s+and\s+(?:shareholders?|stockholders?|owners?)"],
    "current_assets": [r"^total\s+current\s+assets$"],
    "cash": [r"^cash\s+and\s+cash\s+equivalents$", r"^cash,?\s+cash\s+equivalents",
             r"^cash\s+and\s+(?:cash\s+)?equivalents",
             r"^cash$",  # DSGX: bare "Cash" label
             r"^cash\s+and\s+bank\s+balances?$"],
    "total_liabilities": [r"^total\s+liabilities$",
                          r"^total\s+liabilities\b"],
    "current_liabilities": [r"^total\s+current\s+liabilities$"],
    "short_term_debt": [r"^short[\s-]*term\s+(?:borrowings?|loans?|debt|bank\s+loans?)",
                        r"^(?:bank\s+)?(?:borrowings?|loans?),?\s+current$",
                        r"^current\s+(?:portion\s+of\s+)?(?:borrowings?|loans?|debt|bank\s+loans?)$",
                        r"^(?:borrowings?|loans?)\s*(?:,\s*current|[-–]\s*current)$"],
    "long_term_debt": [r"^long[\s-]*term\s+(?:debt|borrowings?|notes|loan|bank)",
                       r"^non[\s-]*current\s+(?:borrowings?|debt)",
                       r"^(?:long[\s-]*term\s+)?(?:bank\s+)?borrowings?,?\s+non[\s-]*current"],
    "total_equity": [r"^total\s+(?:shareholders?|stockholders?|owners?)[\u2019']?\s*equity$",
                     r"^total\s+equity$",
                     r"^total\s+(?:shareholders?|stockholders?).*?equity",
                     r"^equity\s+attributable\s+to.*?(?:shareholders?|stockholders?)",
                     r"^shareholders?[\u2019']?s?\s*equity\b",  # MOMO: "Shareholder's equity (ii)"
],
    "retained_earnings": [r"^retained\s+earnings?\s*\(?accumulated\s+deficit\)?",
                          r"^retained\s+(?:earnings?|profits?)",
                          r"^accumulated\s+(?:deficit|earnings?)",
                          r"^(?:accumulated\s+)?(?:deficit|surplus)"],
    "short_term_investments": [r"^short[\s-]*term\s+investments?", r"^(?:current\s+)?marketable\s+securities"],
    "accounts_receivable": [r"^(?:accounts?|trade)\s+receivable"],
    "inventory": [r"^inventor(?:y|ies)$"],
    "goodwill": [r"^goodwill$"],
    "total_non_current_assets": [r"^total\s+non[\s-]*current\s+assets$"],
    "accounts_payable": [r"^(?:accounts?|trade)\s+payable"],
    "deferred_revenue": [r"^deferred\s+(?:revenues?|income)$",
                         r"^(?:customer\s+advances?\s+and\s+)?deferred\s+revenues?$"],
    "total_non_current_liabilities": [r"^total\s+non[\s-]*current\s+liabilities$"],
    "additional_paid_in_capital": [r"^additional\s+paid[\s-]*in\s+capital$"],
}
CF_PATTERNS = {
    "operating_cf": [r"^(?:net\s+)?cash\s+(?:provided|generated|used)\s+(?:by|in|from)\s+operating",
                     r"^net\s+cash\s+(?:from|used\s+in|used\s+for)\s+operating",
                     r"^(?:net\s+)?cash\s+(?:flows?\s+)?(?:from|provided\s+by|generated\s+from)\s+operating",
                     r"^net\s+cash\s+provided\s+by\s+\(?used\s+in\)?\s+operating",
                     r"^(?:net\s+)?cash.*operating\s+activities"],   # 宽泛匹配 Apple 风格
    "investing_cf": [r"^(?:net\s+)?cash\s+(?:provided|generated|used)\s+(?:by|in|from)\s+investing",
                     r"^net\s+cash\s+(?:from|used\s+in|used\s+for)\s+investing",
                     r"^(?:net\s+)?cash\s+(?:flows?\s+)?(?:from|used\s+in)\s+investing",
                     r"^net\s+cash\s+provided\s+by\s+\(used\s+for\)\s+investing",
                     r"^net\s+cash\s+(?:provided\s+by\s+)?\(?used\s+in\)?\s+investing",
                     r"^(?:net\s+)?cash.*investing\s+activities"],   # 宽泛匹配 Apple 风格
    "financing_cf": [r"^(?:net\s+)?cash\s+(?:provided|generated|used)\s+(?:by|in|from)\s+financing",
                     r"^net\s+cash\s+(?:from|used\s+in|used\s+for)\s+financing",
                     r"^(?:net\s+)?cash\s+(?:flows?\s+)?(?:from|used\s+in)\s+financing",
                     r"^net\s+cash\s+(?:provided|used).*?financing",
                     r"^net\s+cash\s+provided\s+by\s+\(?used\s+in\)?\s+financing",
                     r"^(?:net\s+)?cash.*financing\s+activities"],   # 宽泛匹配 Apple 风格
    "capex": [r"^(?:purchases?\s+of|payments?\s+for|additions?\s+to)\s+property",
              r"^(?:purchases?\s+of|payments?\s+for|additions?\s+to).*property",   # Apple: "Payments for acquisition of property..."
              r"^capital\s+expenditure",
              r"^purchase\s+of\s+property\s+and\s+equipment",
              r"^(?:purchases?\s+of|payments?\s+for)\s+(?:fixed|tangible)\s+assets"],
    "depreciation_amortization": [r"^depreciation\s+and\s+amortization", r"^depreciation"],
    "stock_based_compensation": [r"^stock[\s-]*based\s+compensation", r"^share[\s-]*based\s+compensation",
                               r"^share\s+based\s+long[\s-]*term\s+incentive"],
    "acquisitions": [r"^acquisitions?(?:,|\s+of|\s+net)"],
    "net_change_in_cash": [r"^(?:net\s+)?(?:increase|decrease)\s+in\s+cash"],
    "cash_end_of_period": [r"^cash.*?(?:end\s+of|at\s+end)"],
}

# ═══════════ 股本(正文后备) ═══════════
def raw_shares_to_m(val):
    """将原始股本数字（不同量级）统一转为 M（百万）单位"""
    if val is None or val <= 0: return None
    if val >= 1e7: return round(val / 1e6, 2)   # 原始股数（≥1千万）→ M（如 21,821,589 → 21.8M）
    if val >= 1e5: return round(val / 1e3, 2)   # 千股单位（≥10万）→ M
    return round(val, 2)                         # 已是 M

def extract_shares_from_text(text):
    info = SharesInfo()
    # 解析 ADS 比例，统一存为"每1 ADS 对应 N 普通股"（N 可为小数）
    # 辅助：英文数字词 → 数值
    _WORD_NUM = {"one":1,"two":2,"three":3,"four":4,"five":5,"six":6,"seven":7,
                 "eight":8,"nine":9,"ten":10,"fifteen":15,"twenty":20,"thirty":30}
    def _parse_num_or_word(s):
        s = s.strip().lower()
        if s in _WORD_NUM: return float(_WORD_NUM[s])
        try: return float(s)
        except: return None
    # 数字或英文数字词的通用模式
    _NUM_PAT = r"(\d+(?:\.\d+)?|one|two|three|four|five|six|seven|eight|nine|ten|fifteen|twenty|thirty)"
    # 搜索范围扩大到前 300K 字（6-K 中 ADS 声明可能较靠后）
    _search_text = text[:300000]

    def _set_ratio(n_ord, n_ads):
        if n_ads and n_ord is not None:
            ratio = round(n_ord / n_ads, 4)
            info.ads_ratio = f"1 ADS = {ratio} shares"
            info.ads_per_share = ratio
            return True
        return False

    # 模式1: "N ordinary shares equals/= M ADS"
    m = re.search(_NUM_PAT + r"\s*(?:ordinary|common)\s*shares?\s*(?:equals?\s+(?:to\s+)?|=)\s*" + _NUM_PAT + r"\s*ADS",
                  _search_text, re.I)
    if m:
        _set_ratio(_parse_num_or_word(m.group(1)), _parse_num_or_word(m.group(2)))
    if not info.ads_per_share:
        # 模式2: "each/N ADS represents/= N ordinary shares"
        m = re.search(r"(?:(?:" + _NUM_PAT + r"\s*)?ADS[s]?|each\s+ADS)\s*(?:represent\w*|=|equal\w*)\s*" + _NUM_PAT
                      + r"\s*(?:\(\d+\)\s*)?(?:of\s+the\s+Company'?s?\s+)?(?:Class\s*\w\s+)?(?:ordinary|common)",
                      _search_text, re.I)
        if m:
            n_ads_raw = m.group(1) if m.group(1) else "1"
            _set_ratio(_parse_num_or_word(m.group(2)), _parse_num_or_word(n_ads_raw))
    if not info.ads_per_share:
        # 模式3: "each representing N ordinary shares"（如 CAN）
        m = re.search(r"each\s+representing\s+" + _NUM_PAT + r"\s*(?:\(\d+\)\s*)?(?:of\s+the\s+\S+\s+)?(?:Class\s*\w\s+)?(?:ordinary|common)",
                      _search_text, re.I)
        if m:
            _set_ratio(_parse_num_or_word(m.group(1)), 1.0)
    if not info.ads_per_share:
        # 模式4: "1 ordinary share equals M ADS"（如 VIPS: 1 ordinary = 5 ADS）
        m = re.search(_NUM_PAT + r"\s*ordinary\s*shares?\s*(?:equals?\s+(?:to\s+)?|=)\s*" + _NUM_PAT + r"\s*ADS",
                      _search_text, re.I)
        if m:
            _set_ratio(_parse_num_or_word(m.group(1)), _parse_num_or_word(m.group(2)))
    if not info.ads_per_share:
        # 模式5: "each ADS represents N share(s) of a common/ordinary share"（如 DDI: 0.05 share of a common share）
        m = re.search(r"each\s+ADS\s*represents?\s*" + _NUM_PAT + r"\s*(?:share\s+of\s+(?:a\s+)?)?(?:ordinary|common)",
                      _search_text, re.I)
        if m:
            _set_ratio(_parse_num_or_word(m.group(1)), 1.0)
    for pat, stype in [(r"([0-9][0-9,]+)\s*(?:Class\s*[A-Z]\s+)?(?:ordinary|common)\s*shares?"
                        r"\s*(?:issued\s*and\s*)?outstanding", "common"),
                       (r"(?:shares?|stock)\s*outstanding[:\s,]+([0-9][0-9,]+)", "common")]:
        for m in re.finditer(pat, text[:50000], re.I):
            val = parse_num(m.group(1))
            if val and val > 1000:
                val_m = raw_shares_to_m(val)
                info.details.append({"type": stype, "value": val_m, "raw_value": val})
                if stype == "common" and info.common is None: info.common = val_m
    return info

# ═══════════ BS 节区求和（用于没有小计行的报表）═══════════
def find_section_subtotal(rows, section_start_pat, col, unit_mult=1.0):
    """找节标题（无数值行）之后第一个无标签的纯数字行，作为小计值返回。
    适用于 LULU 式报表：'Current assets' 后跟明细行，小计行无标签，之后直接是非流动资产明细。
    """
    in_section = False
    for row in rows:
        label = build_row_label(row).lower().strip()
        nums = extract_numbers_from_row(row)
        if not in_section:
            if re.search(section_start_pat, label, re.I) and not nums:
                in_section = True
            continue
        # 第一个无标签、有数值的行就是小计行
        if not label and nums:
            v = nums[col] if len(nums) > col else nums[0]
            if v and v > 0:
                return round(v * unit_mult, 2)
        # 如果遇到有标签且有数值的行，说明小计在后面还没出现；继续
        # 如果遇到有标签且无数值（新节标题），不应该出现在此处，继续
    return None


def sum_bs_section(rows, section_start_pat, section_end_pat, col, unit_mult):
    """在 BS 表中找到 section_start_pat 节标题后，累加各明细行数值直到 section_end_pat。
    - 起始行本身必须无数值（纯节标题）
    - 跳过空标签行（小计/合计行避免重复计算）
    """
    in_section = False
    total = 0.0
    has_items = False
    for row in rows:
        label = build_row_label(row).lower().strip()
        nums = extract_numbers_from_row(row)
        if re.search(section_start_pat, label) and not nums:
            in_section = True
            continue
        if in_section and label and re.search(section_end_pat, label):
            break
        if in_section and nums and label:   # 跳过空标签行（小计行）
            v = nums[col] if len(nums) > col else nums[0]
            if v is not None:
                total += abs(v)
                has_items = True
    return round(total * unit_mult, 2) if has_items else None


# ═══════════ 主分析 ═══════════
def analyze_filing(filepath, ticker="", _silent=False):
    if not _silent:
        print(f"  [parse] {filepath.name} ...", end=" ", flush=True)
    raw = read_filing(filepath)
    html = extract_html_from_sgml(raw)
    soup = BeautifulSoup(html, "lxml")
    all_tables = soup.find_all("table")
    text = soup.get_text(separator="\n", strip=True)
    if len(text) < 500:
        if not _silent: print("(too short)")
        return FilingData(ticker=ticker, file_name=filepath.name, filing_type="Unknown")

    d = FilingData()
    d.ticker = ticker or filepath.parent.name
    d.file_name = filepath.name
    d.filing_type = detect_filing_type(html, filepath.name)
    d.company_name = extract_company_name_sgml(raw) or extract_company_name(text, ticker)
    d.filing_date, d.fiscal_period = extract_date_info(text)
    # 日期修正：
    # htm/html (XBRL) 文件 → 正文日期是文件内部签署日期，SEC 提交日期在文件名中，以文件名为准
    # txt (SGML) 文件 → 正文或 SGML header 中的日期通常准确，仅在缺失时用文件名回退
    fn_date = date_from_filename(filepath.name)
    if filepath.suffix.lower() in (".htm", ".html", ".xml"):
        if fn_date:
            d.filing_date = fn_date   # XBRL 文件始终用文件名日期
    elif not d.filing_date and fn_date:
        d.filing_date = fn_date       # SGML 文件仅在缺失时回退
    d.currency = detect_currency(text)

    # Step 1: 分类表格
    classified = {"IS": [], "BS": [], "CF": []}
    for i, table in enumerate(all_tables):
        cat = classify_table(table)
        if cat:
            rows = parse_table_to_rows(table)
            ul, um = detect_table_unit(table, html)
            classified[cat].append({"index": i, "rows": rows, "row_count": len(rows),
                                    "unit_label": ul, "unit_mult": um})

    def _unit_reliability(ul):
        """Score unit detection reliability: intrinsic > nearby/dom > doc-fallback > unknown.
        Tables with explicit headers/xbrl/notation are always preferred over doc-fallback.
        'ones' (Method 4 heuristic) is slightly less reliable than explicit declarations
        to avoid false positives where a thousands-unit table is misclassified as ones."""
        if ul in ("thousands", "millions", "billions",
                  "thousands(xbrl)", "millions(xbrl)", "ones(xbrl)", "billions(xbrl)",
                  "thousands(notation)"):
            return 3  # intrinsic declaration
        if ul in ("ones",):
            return 2  # Method 4 heuristic: reliable but can false-positive on thousands tables
        if ul in ("thousands(nearby)", "millions(nearby)", "thousands(nearby-after)",
                  "millions(nearby-after)", "thousands(dom)", "millions(dom)",
                  "billions(dom)", "ones(expressed-currency)"):
            return 1  # contextual inference
        return 0      # doc-fallback or unknown

    def pick(lst):
        """Pick the best table: prefer explicit unit declarations, then most rows."""
        if not lst: return None
        best_rel = max(_unit_reliability(t["unit_label"]) for t in lst)
        reliable = [t for t in lst if _unit_reliability(t["unit_label"]) == best_rel]
        return max(reliable, key=lambda t: t["row_count"])

    def pick_is(lst):
        """IS 选表：优先选含 'revenue' 关键词的 GAAP 利润表，排除非 GAAP 调节表"""
        if not lst: return None
        def has_revenue(t):
            top = " ".join(" ".join(r) for r in t["rows"][:8]).lower()
            if re.search(r'\brevenues?\b', top): return True
            # Also check if any row in the first 8 rows matches a revenue pattern (e.g. MRK: bare "Sales")
            for row in t["rows"][:8]:
                lbl = build_row_label(row).lower().strip()
                if lbl and any(re.match(p, lbl, re.I) for p in IS_PATTERNS["revenue"]):
                    return True
            # Bank/financial IS tables: revenue appears deeper (after interest income section)
            top30 = " ".join(" ".join(r) for r in t["rows"][:30]).lower()
            return bool(re.search(r'\btotal\s+revenue\b|net\s+interest\s+income', top30))
        def is_nongaap(t):
            full = " ".join(" ".join(r) for r in t["rows"]).lower()
            return full.count("non-gaap") >= 3
        def is_reconciliation(t):
            """Detect GAAP/IFRS reconciliation adjustment tables — not the primary P&L."""
            full = " ".join(" ".join(r) for r in t["rows"][:10]).lower()
            return bool(re.search(r'ifrs\s+adjustment|gaap\s+reconciliation|reconcil.*?to.*?gaap|as\s+reported\s+under\s+(?:us\s+)?gaap', full))
        def is_segment_table(t):
            """Detect segment disclosure tables — multiple segment columns, not a primary P&L."""
            # Segment tables have "revenue from external/intersegment customers" row labels
            top = " ".join(" ".join(r) for r in t["rows"][:8]).lower()
            if re.search(r'revenue\s+from\s+(?:external|intersegment)', top):
                return True
            # Segment tables have non-temporal column headers (geography/business-unit names).
            # Multi-period P&Ls have only temporal headers (years, month-end dates).
            # Check if any header row has >= 3 non-date column cells (cols 1+) with text.
            # Temporal cells contain "month", "quarter", "ended", "period", or bare 4-digit years.
            _temporal_pat = re.compile(
                r'\b(?:month|quarter|ended|year|period|20\d{2}|19\d{2}|january|february|march'
                r'|april|may|june|july|august|september|october|november|december)\b', re.I)
            if t["rows"]:
                for row in t["rows"][:6]:
                    nontemporal = sum(
                        1 for i, c in enumerate(row)
                        if i > 0 and c.strip() and re.search(r'[a-zA-Z]', c)
                        and not _temporal_pat.search(c)
                    )
                    if nontemporal >= 3:
                        return True
            return False
        def is_field_count(t):
            """Count how many IS_PATTERNS fields are matched in this table (quality metric)."""
            count = 0
            for pats in IS_PATTERNS.values():
                for row in t["rows"]:
                    lbl = build_row_label(row).lower().strip()
                    if lbl and any(re.match(p, lbl, re.I) for p in pats):
                        count += 1
                        break
            return count
        gaap = [t for t in lst if not is_nongaap(t) and not is_reconciliation(t) and not is_segment_table(t)]
        candidates = gaap if gaap else [t for t in lst if not is_nongaap(t) and not is_reconciliation(t)]
        candidates = candidates if candidates else [t for t in lst if not is_nongaap(t)]
        candidates = candidates if candidates else lst
        with_rev = [t for t in candidates if has_revenue(t)]
        candidates = with_rev if with_rev else candidates
        # Prefer tables with explicit unit declarations over doc-fallback/unknown
        best_rel = max(_unit_reliability(t["unit_label"]) for t in candidates)
        reliable = [t for t in candidates if _unit_reliability(t["unit_label"]) == best_rel]
        candidates = reliable if reliable else candidates
        # Score by (field_matches, row_count) — prefers tables matching more IS fields
        return max(candidates, key=lambda t: (is_field_count(t), t["row_count"]))
    best_is, best_bs, best_cf = pick_is(classified["IS"]), pick(classified["BS"]), pick(classified["CF"])
    if not _silent:
        print(f"[IS={len(classified['IS'])} BS={len(classified['BS'])} CF={len(classified['CF'])}]", end=" ")

    # Step 2: 全局单位（优先取可靠性最高的表）
    gu, gl = 0.0, "unknown"
    _best_gu_rel = -1
    for t in [best_is, best_bs, best_cf]:  # priority: best tables first
        if t and t["unit_mult"] > 0:
            r = _unit_reliability(t["unit_label"])
            if r > _best_gu_rel:
                gu, gl, _best_gu_rel = t["unit_mult"], t["unit_label"], r
    # Also scan ALL classified tables — sometimes best_is has a low-reliability unit
    # (e.g. 'ones' from Method 4) while another IS/BS table has 'thousands(xbrl)' (rel=3).
    for cat_tables in classified.values():
        for t in cat_tables:
            if t["unit_mult"] > 0:
                r = _unit_reliability(t["unit_label"])
                if r > _best_gu_rel:
                    gu, gl, _best_gu_rel = t["unit_mult"], t["unit_label"], r

    # Per-ticker unit override: for filings that omit explicit "in thousands/millions"
    # IMPP says "Expressed in United States Dollars" = ones (individual USD)
    _KNOWN_UNIT_OVERRIDES: Dict[str, float] = {
        "IMPP":  1e-6,   # Imperial Petroleum: "Expressed in USD" = individual dollars (ones)
        "IMPPP": 1e-6,   # Same company preferred shares
    }
    # Document-level "expressed/presented in [currency]" detection
    # When a filing declares currency without thousands/millions scale (e.g. "Expressed in United States Dollars",
    # "Presented in Euros"), values are individual currency units (ones). Set gu=1e-6 to convert to M.
    if gu == 0:
        _doc_up = re.sub(r"<[^>]+>", " ", html).upper()
        _has_scale = bool(re.search(r"IN\s+THOUSANDS|IN\s+MILLIONS|IN\s+BILLIONS", _doc_up))
        if not _has_scale:
            _currency_pat = (r"(?:EXPRESSED|PRESENTED|STATED|DENOMINATED)\s+IN\s+"
                             r"(?:UNITED\s+STATES\s+DOLLARS|US\s+DOLLARS|EUROS?|EUR\b|"
                             r"BRITISH\s+POUNDS?|GBP\b|CANADIAN\s+DOLLARS?|CAD\b|"
                             r"NORWEGIAN\s+KRONE|NOK\b|SWEDISH\s+KRONOR|SEK\b|"
                             r"SWISS\s+FRANCS?|CHF\b|JAPANESE\s+YEN|JPY\b)")
            # Exclude "CURRENCY EXPRESSED IN ..." — that phrase just names the currency,
            # not the scale (e.g. MWG: "Currency expressed in USD" but values are in $'000).
            _has_currency_qualifier = bool(re.search(r"CURRENCY\s+EXPRESSED\s+IN", _doc_up))
            if re.search(_currency_pat, _doc_up) and not _has_currency_qualifier:
                gu, gl = 1e-6, "ones(expressed-currency)"
            # Also catch "(in USD)" / "(in U.S. dollars)" parenthetical — common in smaller filings
            elif re.search(r"\(\s*IN\s+(?:USD|U\.S\.\s+DOLLARS?|US\s+DOLLARS?)\s*[,)]", _doc_up):
                gu, gl = 1e-6, "ones(expressed-currency)"
    if gu == 0:
        _uoverride = _KNOWN_UNIT_OVERRIDES.get(d.ticker.upper())
        if _uoverride:
            if abs(_uoverride - 1e-6) < 1e-12: _ulabel = "ones(override)"
            elif abs(_uoverride - 0.001) < 1e-9: _ulabel = "thousands(override)"
            else: _ulabel = "millions(override)"
            gu, gl = _uoverride, _ulabel
    def _um(ti):
        if not ti or ti["unit_mult"] <= 0: return gu
        # Fall back to global unit when it's more reliable (e.g. table=ones heuristic,
        # global=thousands(xbrl) from another table in the same document).
        if gu > 0 and _unit_reliability(ti["unit_label"]) < _unit_reliability(gl):
            return gu
        return ti["unit_mult"]
    def ap(val, ti, is_eps=False):
        if val is None: return None
        if is_eps: return val
        um = _um(ti)
        return round(val * um, 2) if um > 0 else val
    # Pre-set d.unit_multiplier = gu so that Step 4 auto-infer triggers when gu==0.
    # Without this, filings with no IS table (best_is=None) keep the default 1.0,
    # preventing auto-infer and leaving raw BS values unscaled (e.g. API: 721099 → 721099M).
    d.unit_multiplier = gu

    # Step 3: 三大报表
    if best_is:
        is_col = detect_latest_column(best_is["rows"])
        isd = extract_from_table(best_is["rows"], IS_PATTERNS, col=is_col)
        d.revenue = ap(isd["revenue"], best_is)
        # Fallback: revenue may be an unlabeled subtotal row under a "Revenues:" section header
        # (e.g. HOLX: "Revenues:" header → Product + Service lines → unlabeled total)
        if d.revenue is None:
            _rev_unlabeled = find_section_subtotal(
                best_is["rows"], r"^revenues?:?$", is_col,
                unit_mult=_um(best_is) if _um(best_is) > 0 else 1.0)
            if _rev_unlabeled:
                d.revenue = _rev_unlabeled
        d.cost_of_revenue = ap(isd["cost_of_revenue"], best_is)
        d.gross_profit = ap(isd["gross_profit"], best_is)
        d.operating_expenses = ap(isd["operating_expenses"], best_is)
        d.operating_income = ap(isd["operating_income"], best_is)
        d.net_income = ap(isd["net_income"], best_is)
        d.eps_basic = ap(isd["eps_basic"], best_is, is_eps=True)
        d.eps_diluted = ap(isd["eps_diluted"], best_is, is_eps=True)
        d.rd_expense = ap(isd.get("rd_expense"), best_is)
        d.sm_expense = ap(isd.get("sm_expense"), best_is)
        d.ga_expense = ap(isd.get("ga_expense"), best_is)
        d.sga_expense = ap(isd.get("sga_expense"), best_is)
        d.interest_income = ap(isd.get("interest_income"), best_is)
        d.interest_expense = ap(isd.get("interest_expense"), best_is)
        d.income_tax = ap(isd.get("income_tax"), best_is)
        d.pretax_income = ap(isd.get("pretax_income"), best_is)
        _eff_um = _um(best_is)
        d.unit_multiplier = _eff_um
        # Use global unit label when _um fell back to gu (table unit was less reliable)
        d.unit_label = gl if (gu > 0 and _eff_um == gu and best_is["unit_label"] != gl) else (best_is["unit_label"] if best_is["unit_mult"] > 0 else gl)
        # Bank: non-interest expense may appear as section header with no aggregate value;
        # sum sub-items between header and next major section.
        # Use raw unit (1.0) when unit_mult=0 so step-4 auto-scaling handles it.
        if d.operating_expenses is None:
            _sum_um = _um(best_is) if _um(best_is) > 0 else 1.0
            v = sum_bs_section(best_is["rows"],
                               r"^non.interest\s+expense$",
                               r"^income\s+before|^provision\s+for|^pre.provision",
                               is_col, _sum_um)
            if v:
                d.operating_expenses = v

    # IS null-field fallback: scan other IS tables for fields still missing from best_is
    _is_null_fields = [f for f in ["gross_profit","operating_expenses","operating_income",
                                    "interest_income","interest_expense","rd_expense",
                                    "sm_expense","ga_expense","sga_expense"] if getattr(d, f) is None]
    if _is_null_fields:
        for is_tbl in classified["IS"]:
            if is_tbl is best_is: continue
            fb_col = detect_latest_column(is_tbl["rows"])
            fb = extract_from_table(is_tbl["rows"], {k: IS_PATTERNS[k] for k in _is_null_fields
                                                      if k in IS_PATTERNS}, col=fb_col)
            for k in list(_is_null_fields):
                if getattr(d, k) is None and fb.get(k) is not None:
                    setattr(d, k, ap(fb[k], is_tbl))
                    _is_null_fields.remove(k)
            if not _is_null_fields: break

    # EPS: 如果最大IS表没拿到，遍历其他IS表
    if d.eps_basic is None or d.eps_diluted is None:
        for is_tbl in classified["IS"]:
            if is_tbl is best_is: continue
            eps_col = detect_latest_column(is_tbl["rows"])
            tmp = extract_from_table(is_tbl["rows"], {
                "eps_basic": IS_PATTERNS["eps_basic"],
                "eps_diluted": IS_PATTERNS["eps_diluted"]}, col=eps_col)
            if d.eps_basic is None and tmp["eps_basic"]: d.eps_basic = tmp["eps_basic"]
            if d.eps_diluted is None and tmp["eps_diluted"]: d.eps_diluted = tmp["eps_diluted"]
            if d.eps_basic and d.eps_diluted: break

    # EPS: context-aware extraction for "-Basic"/"-Diluted" sub-rows (PDD style)
    if d.eps_basic is None or d.eps_diluted is None:
        for is_tbl in classified["IS"]:
            ec_col = detect_latest_column(is_tbl["rows"])
            eb, ed = extract_eps_contextual(is_tbl["rows"], ec_col)
            if d.eps_basic is None and eb: d.eps_basic = eb
            if d.eps_diluted is None and ed: d.eps_diluted = ed
            if d.eps_basic and d.eps_diluted: break

    # EPS 健全性检查：|EPS| > 500 说明误提取到股本数或其他大数，丢弃并重试上下文感知提取
    def _eps_sane(v): return v is not None and abs(v) <= 500
    if not _eps_sane(d.eps_basic) or not _eps_sane(d.eps_diluted):
        d.eps_basic = d.eps_basic if _eps_sane(d.eps_basic) else None
        d.eps_diluted = d.eps_diluted if _eps_sane(d.eps_diluted) else None
        for is_tbl in classified["IS"]:
            ec_col = detect_latest_column(is_tbl["rows"])
            eb, ed = extract_eps_contextual(is_tbl["rows"], ec_col)
            if d.eps_basic is None and _eps_sane(eb): d.eps_basic = eb
            if d.eps_diluted is None and _eps_sane(ed): d.eps_diluted = ed
            if d.eps_basic and d.eps_diluted: break

    # 股本：先上下文感知提取（-Basic/-Diluted），再标准模式后备
    def shares_to_m(v, in_thousands=None):
        """将 IS 表中的股本数转为 M（百万）单位。
        in_thousands=True  → 表单位为 thousands，v 除以 1000 得 M
        in_thousands=False → 表单位为 millions，v 已是 M，直接返回（修复：之前误除1000）
        in_thousands=None  → 自动推断
        """
        if v is None or v <= 0: return None
        if in_thousands is True:
            result = round(v / 1000, 2)
            # 健全性：若 result > 30B，v 极可能是原始股数（raw count），而非 thousands 单位
            # 例如 IH: 54,011,420 raw ADSs in a thousands table → /1000=54011M(wrong) → /1e6=54M(✓)
            if result > 30_000:
                return round(v / 1e6, 2)
            return result
        elif in_thousands is False:
            # 股数是原始计数（raw count），如 "in thousands except share data"
            # v >= 1e6 → 除以 1e6 转为 M；否则视为已是 M
            if v >= 1e6:
                return round(v / 1e6, 2)
            return round(v, 2)   # 已是 M（如中国公司以百万为单位列示股数）
        else:
            if v >= 1e7: return round(v / 1e6, 2)
            if v > 10000: return round(v / 1000, 2)
            return round(v, 2)

    # Process IS tables ordered by unit reliability (highest first) so that XBRL-annotated
    # or explicit-header tables take priority over Method 4 'ones' heuristic tables.
    # This prevents ones-unit tables (reliability=2) from overriding correctly-detected
    # thousands/millions tables (reliability=3) in the shares extraction loop.
    _is_ordered = sorted(classified["IS"], key=lambda t: _unit_reliability(t["unit_label"]), reverse=True)
    for is_tbl in _is_ordered:
        sc_col = detect_latest_column(is_tbl["rows"])
        tbl_txt = " ".join(" ".join(r) for r in is_tbl["rows"][:20])
        sb, sd, sh_ik, _ads_flag = extract_shares_contextual(is_tbl["rows"], sc_col, tbl_txt)
        _tbl_um = is_tbl.get("unit_mult", 0)
        # Override sh_ik using the table's detected unit_mult when the table text itself
        # does not contain the unit declaration (e.g. NTES: "thousands" is in nearby HTML).
        # Respect "except share/per-share" clauses which mean shares are NOT in thousands.
        if _tbl_um > 0 and not re.search(r"except\s+(?:for\s+)?(?:share|per[\s\-]share)", tbl_txt, re.I):
            if abs(_tbl_um - 0.001) < 1e-9: sh_ik = True   # thousands table
            elif abs(_tbl_um - 1.0) < 1e-9: sh_ik = False  # millions table
        # Additional positive check: filing explicitly says shares are reflected in thousands
        # (e.g. AAPL: "in millions, except number of shares which are reflected in thousands")
        # The "except number of shares" pattern doesn't match the regex above, so re-check here.
        if re.search(r"reflected\s+in\s+thousands", tbl_txt, re.I):
            sh_ik = True
        def _sh_m(v):
            if not v or v <= 100: return None
            # "ones" 表（unit_mult ≈ 1e-6）：原始股数直接 /1e6 → M
            if abs(_tbl_um - 1e-6) < 1e-9:
                return round(v / 1e6, 2)
            return shares_to_m(v, sh_ik)
        # In a millions-unit table with sh_ik=False, raw share values >= 1e6 are suspicious:
        # the shares are likely in thousands within a mixed-unit table (e.g. AAPL 10-K/10-Q where
        # XBRL decimals=-6 wins majority but share rows have decimals=-3).
        # Skip storing these — the SHARES_PATTERNS fallback will find a dedicated thousands table.
        _raw_sh = max(sb or 0, sd or 0)
        _suspicious = (abs(_tbl_um - 1.0) < 1e-9 and sh_ik != True and _raw_sh >= 1e6)
        if not _suspicious:
            if sb and sb > 100: d.shares.weighted_avg = _sh_m(sb)
            if sd and sd > 100: d.shares.diluted = _sh_m(sd)
            if _ads_flag: d.shares.wtdavg_already_ads = True
            if d.shares.weighted_avg or d.shares.diluted:
                break
    if not d.shares.weighted_avg and not d.shares.diluted:
        # Build expanded table list: IS + BS + any table with "weighted"/"shares".
        # Exclude CF tables: cash flow statements contain items like "Issuance of ordinary shares: $25,000"
        # which SHARES_PATTERNS can falsely match, producing dollar amounts as share counts.
        _search = classified["IS"][:] + classified["BS"][:]
        for ii, table in enumerate(all_tables):
            txt = table.get_text(separator=" ", strip=True).lower()
            if ("weighted" in txt and "share" in txt) or "earnings per share" in txt:
                if not any(t["index"] == ii for t in _search):
                    rr = parse_table_to_rows(table)
                    ul2, um2 = detect_table_unit(table, html)
                    _search.append({"index": ii, "rows": rr, "row_count": len(rr), "unit_label": ul2, "unit_mult": um2})
        for stbl in _search:
            # Skip anti-dilutive disclosure tables: they report the number of shares EXCLUDED
            # from diluted EPS (not the total share count). e.g. NUE: anti-dilutive table has
            # "Weighted-average shares: 164" (164K excluded shares, not 164M total shares).
            _antidilut_check = " ".join(" ".join(r) for r in stbl["rows"][:5]).lower()
            if re.search(r'anti.?dilut', _antidilut_check):
                continue
            sc = detect_latest_column(stbl["rows"])
            tt = " ".join(" ".join(r) for r in stbl["rows"][:20])
            _stbl_um = stbl.get("unit_mult", 0)
            _has_except = bool(re.search(r"except\s+(?:for\s+)?(?:share|per[\s\-]share)", tt, re.I))
            if _stbl_um > 0:
                if _has_except:
                    ik = False  # "except share data" → shares are raw counts, not scaled by table unit
                elif abs(_stbl_um - 0.001) < 1e-9:
                    ik = True   # thousands table, no exception → shares also in thousands
                elif abs(_stbl_um - 1.0) < 1e-9:
                    ik = False  # millions table
                else:
                    ik = "thousand" in tt.lower()
            else:
                ik = "thousand" in tt.lower()
            # Same positive check: shares reflected in thousands within a millions table
            if re.search(r"reflected\s+in\s+thousands", tt, re.I):
                ik = True
            shd = extract_from_table(stbl["rows"], SHARES_PATTERNS, col=sc)
            sb, sd = shd.get("shares_basic"), shd.get("shares_diluted")
            # Use result-based threshold (>= 0.1M) instead of raw value > 100,
            # so that millions-unit tables where shares appear as small numbers (e.g. 32.4)
            # are correctly stored (e.g. BKNG: 32.4M in millions table → raw 32.4 ≤ 100 but result=32.4M).
            _sb_m = shares_to_m(sb, ik) if (sb and sb > 0) else None
            _sd_m = shares_to_m(sd, ik) if (sd and sd > 0) else None
            if _sb_m and _sb_m >= 0.1: d.shares.weighted_avg = _sb_m
            if _sd_m and _sd_m >= 0.1: d.shares.diluted = _sd_m
            # 检测 IS 是否以 ADS 为单位报告加权平均
            if re.search(r'\bads[s]?\b', tt, re.I): d.shares.wtdavg_already_ads = True
            if d.shares.weighted_avg or d.shares.diluted: break

    if d.shares.diluted and not d.shares.weighted_avg:
        d.shares.weighted_avg = d.shares.diluted

    # Sanity: diluted shares must be >= weighted_avg (basic). If diluted is dramatically smaller
    # than weighted_avg, the "diluted" value likely came from a dilutive-effect row (incremental
    # shares from options/RSUs) rather than the total diluted shares count.
    # E.g. BKNG: weighted_avg=32.45M (correct basic) but diluted=0.187M (dilutive effect only).
    if d.shares.diluted and d.shares.weighted_avg:
        if d.shares.diluted < d.shares.weighted_avg * 0.5:
            d.shares.diluted = d.shares.weighted_avg  # replace bad diluted with basic count

    # Also propagate weighted_avg → diluted so EPS cross-val (which checks diluted) can fire
    # when only weighted_avg was found (e.g. HCA: basic row found but diluted row absent).
    if d.shares.weighted_avg and not d.shares.diluted:
        d.shares.diluted = d.shares.weighted_avg

    # EPS cross-validation: if eps and net_income are available, verify share scale.
    # Detects cases where shares are off by 1000x due to missing "except share data" clause
    # in a thousands-unit table (e.g. BRIA: raw=23342466 shares in USD'000 table → /1000 wrong).
    _eps_for_xval = d.eps_diluted or d.eps_basic  # use diluted, fall back to basic
    if _eps_for_xval and d.net_income and d.shares.diluted:
        _eps_check = abs(_eps_for_xval)
        _ni_check = abs(d.net_income)
        _sh_check = d.shares.diluted
        if _eps_check > 0 and _sh_check > 0:
            _eps_computed = _ni_check / _sh_check
            _ratio = _eps_computed / _eps_check
            if _ratio < 0.005:    # shares ~200x too large → /1000
                _f = round(1 / _ratio)
                if abs(_f - 1000) < 200:  # only correct if ratio is close to 1000
                    d.shares.diluted = round(d.shares.diluted / 1000, 2)
                    if d.shares.weighted_avg:
                        d.shares.weighted_avg = round(d.shares.weighted_avg / 1000, 2)
            elif _ratio > 200:    # shares ~200x too small → *1000
                _f = round(_ratio)
                if abs(_f - 1000) < 200:
                    d.shares.diluted = round(d.shares.diluted * 1000, 2)
                    if d.shares.weighted_avg:
                        d.shares.weighted_avg = round(d.shares.weighted_avg * 1000, 2)

    # BS: merge ALL BS tables (some filings split Assets / Liabilities into separate tables)
    # Process best_bs first so its values have priority over secondary/notes tables
    if classified["BS"]:
        bs_order = ([best_bs] if best_bs else []) + [t for t in classified["BS"] if t is not best_bs]
        for bs_tbl in bs_order:
            bs_col = detect_latest_column(bs_tbl["rows"])
            bsd = extract_from_table(bs_tbl["rows"], BS_PATTERNS, col=bs_col)
            # Scale consistency check for secondary BS tables: if this table's total_assets
            # disagrees with the already-stored value by >10x, its unit detection is wrong
            # (e.g. HCM: best_bs=thousands but secondary=millions(doc) with SAME raw values →
            # secondary gives 1000x inflated current_assets, equity, etc.).
            # Apply a correction factor so the secondary table's values are still usable.
            _bs_scale_correction = 1.0
            if bs_tbl is not best_bs and d.total_assets and bsd.get("total_assets") is not None:
                _sec_ta = ap(bsd["total_assets"], bs_tbl)
                if _sec_ta and _sec_ta > 0 and d.total_assets > 0:
                    _scale_ratio = _sec_ta / d.total_assets
                    if _scale_ratio > 10 or _scale_ratio < 0.1:
                        _bs_scale_correction = d.total_assets / _sec_ta  # rescale to match best_bs
            for k in BS_PATTERNS:
                if getattr(d, k) is None and bsd[k] is not None:
                    _raw_val = ap(bsd[k], bs_tbl)
                    if _raw_val is not None:
                        setattr(d, k, round(_raw_val * _bs_scale_correction, 2))

        # Sanity: reject total_assets from secondary tables if it's less than current_assets
        # (physically impossible — means the secondary table is a notes/subsidiary table).
        if d.total_assets is not None and d.current_assets is not None:
            if d.total_assets < d.current_assets:
                d.total_assets = None

        # Fallback: some filings (e.g. MRK) have "Total Assets" as an unlabeled row in best_bs.
        # Detect by finding the last unlabeled row in the assets section with value > current_assets.
        if d.total_assets is None and d.current_assets is not None and best_bs:
            _rows_ta = best_bs["rows"]
            _um_ta = best_bs["unit_mult"] if best_bs["unit_mult"] > 0 else 1.0
            _col_ta = detect_latest_column(_rows_ta)
            _in_assets = False
            _unlabeled_ta = None
            for _row_ta in _rows_ta:
                _lbl_ta = build_row_label(_row_ta).lower().strip()
                _nums_ta = extract_numbers_from_row(_row_ta)
                if re.search(r"^assets?$", _lbl_ta, re.I) and not _nums_ta:
                    _in_assets = True
                    continue
                if _in_assets and _lbl_ta and re.search(r"^liabilit", _lbl_ta, re.I):
                    break
                if _in_assets and not _lbl_ta and _nums_ta:
                    _v_ta = _nums_ta[_col_ta] if len(_nums_ta) > _col_ta else _nums_ta[0]
                    if _v_ta is not None and _v_ta > 0 and _v_ta * _um_ta > d.current_assets:
                        _unlabeled_ta = round(_v_ta * _um_ta, 2)
            if _unlabeled_ta:
                d.total_assets = _unlabeled_ta

    # CF: merge ALL CF tables (some filings split operating/investing/financing into separate tables)
    if classified["CF"]:
        for cf_tbl in classified["CF"]:
            cf_col = detect_latest_column(cf_tbl["rows"])
            cfd = extract_from_table(cf_tbl["rows"], CF_PATTERNS, col=cf_col)
            for k in CF_PATTERNS:
                if getattr(d, k) is None:
                    val = ap(cfd[k], cf_tbl)
                    if val is not None:
                        setattr(d, k, val)

    # Step 4: 自动推断单位
    if d.unit_multiplier == 0:
        vals = [v for v in [d.revenue, d.total_assets, d.net_income, d.operating_cf] if v and v > 0]
        if vals:
            # Use max: if ANY reference value > 1e5, values are likely in thousands (not millions).
            # Median was too conservative — e.g. a company with revenue=38K but assets=721K
            # would have median 38K < 1e5, giving mult=1.0 (treating 721K as 721B).
            mult = 0.001 if max(vals) > 1e5 else 1.0
            d.unit_multiplier = mult
            d.unit_label = "auto(thousands->M)" if mult == 0.001 else "auto(=M)"
            for attr in ["revenue","cost_of_revenue","gross_profit","operating_expenses",
                         "operating_income","net_income","total_assets","current_assets",
                         "cash","total_liabilities","current_liabilities","short_term_debt","long_term_debt",
                         "total_equity","retained_earnings","operating_cf","investing_cf",
                         "financing_cf","capex",
                         "rd_expense","sm_expense","ga_expense","sga_expense",
                         "interest_income","interest_expense","income_tax","pretax_income",
                         "short_term_investments","accounts_receivable","inventory","goodwill",
                         "total_non_current_assets","accounts_payable","deferred_revenue",
                         "total_non_current_liabilities","additional_paid_in_capital",
                         "depreciation_amortization","stock_based_compensation",
                         "net_change_in_cash","cash_end_of_period","acquisitions"]:
                v = getattr(d, attr)
                if v is not None: setattr(d, attr, round(v * mult, 2))

    # Step 5: 衍生
    if d.gross_profit is None and d.revenue and d.cost_of_revenue:
        d.gross_profit = round(d.revenue - abs(d.cost_of_revenue), 2)
    if d.free_cash_flow is None and d.operating_cf is not None and d.capex is not None:
        d.free_cash_flow = round(d.operating_cf - abs(d.capex), 2)
    # EBITDA = Operating Income + D&A (simplified)
    if d.ebitda is None and d.operating_income is not None and d.depreciation_amortization is not None:
        d.ebitda = round(d.operating_income + abs(d.depreciation_amortization), 2)
    # SGA = S&M + G&A if not directly available
    if d.sga_expense is None and d.sm_expense and d.ga_expense:
        d.sga_expense = round((abs(d.sm_expense) + abs(d.ga_expense)), 2)
        if d.sm_expense < 0: d.sga_expense = -d.sga_expense

    # ── 推导1：所得税 = 税前利润 - 净利润 ──────────────────────────────────
    if d.income_tax is None and d.pretax_income is not None and d.net_income is not None:
        derived_tax = round(d.net_income - d.pretax_income, 2)
        # 合理性检查：|派生税| 不应超过税前利润绝对值（允许少量负税/退税）
        if d.pretax_income != 0 and abs(derived_tax) <= abs(d.pretax_income) * 1.5:
            d.income_tax = derived_tax

    # ── 推导2：营业利润 = 毛利润 - 经营性费用合计 ────────────────────────────
    if d.operating_income is None and d.gross_profit is not None:
        op_items = [d.rd_expense, d.sm_expense, d.ga_expense, d.sga_expense]
        avail = [v for v in op_items if v is not None]
        # 只有 S&M 或 G&A 之一时不足以重建全部费用，跳过；SGA 已含两者则可用单项
        has_sga = d.sga_expense is not None
        has_sm_and_ga = d.sm_expense is not None and d.ga_expense is not None
        if avail and (has_sga or has_sm_and_ga):
            # 去重：若 sga 与 sm+ga 同时存在，用 sga（已合并，避免重复）
            if has_sga and has_sm_and_ga:
                expense_sum = abs(d.sga_expense) + (abs(d.rd_expense) if d.rd_expense else 0)
            else:
                expense_sum = sum(abs(v) for v in avail)
            d.operating_income = round(d.gross_profit - expense_sum, 2)

    # ── 推导2b：银行/无COGS公司：营业利润 = 营收 - 营业费用合计 ──────────────
    if d.operating_income is None and d.revenue is not None and d.operating_expenses is not None:
        if d.cost_of_revenue is None:  # 银行无直接成本行
            derived = round(d.revenue - abs(d.operating_expenses), 2)
            if 0 < derived < d.revenue:
                d.operating_income = derived

    # ── 推导3：流动资产/流动负债合计（BS 无小计行时逐项求和）────────────────
    if (d.current_assets is None or d.current_liabilities is None) and classified["BS"]:
        for bs_tbl in classified["BS"]:
            bs_col = detect_latest_column(bs_tbl["rows"])
            um = _um(bs_tbl)
            if d.current_assets is None:
                # 优先：找节标题后第一个无标签纯数字行（如 LULU 式报表）
                v = find_section_subtotal(bs_tbl["rows"], r"^current\s+assets?$", bs_col, um)
                # 回退：逐项求和（适用于无小计行的报表）
                if v is None:
                    v = sum_bs_section(bs_tbl["rows"], r"^current\s+assets?$",
                                       r"^non[\s-]*current\s+assets?|^total\s+assets", bs_col, um)
                # 合理性保护：流动资产不能超过总资产
                if v and d.total_assets and v > d.total_assets:
                    v = None
                if v: d.current_assets = v
            if d.current_liabilities is None:
                v = find_section_subtotal(bs_tbl["rows"], r"^current\s+liabilit", bs_col, um)
                if v is None:
                    v = sum_bs_section(bs_tbl["rows"], r"^current\s+liabilit",
                                       r"^non[\s-]*current\s+liabilit|^total\s+liabilit", bs_col, um)
                if v and d.total_liabilities and v > d.total_liabilities:
                    v = None
                if v: d.current_liabilities = v

    # ── 推导3b：总资产/总负债（BS 无明确合计行时逐项求和）────────────────────
    if (d.total_assets is None or d.total_liabilities is None) and classified["BS"]:
        for bs_tbl in classified["BS"]:
            bs_col = detect_latest_column(bs_tbl["rows"])
            _sum_um = _um(bs_tbl) if _um(bs_tbl) > 0 else 1.0
            if d.total_assets is None:
                v = sum_bs_section(bs_tbl["rows"], r"^assets?$",
                                   r"^liabilit", bs_col, _sum_um)
                if v: d.total_assets = v
            if d.total_liabilities is None:
                v = sum_bs_section(bs_tbl["rows"],
                                   r"^liabilities\s+and\s+shareholders|^liabilities\s+and\s+stockholders|^liabilities$",
                                   r"^shareholders.?\s+equity|^stockholders.?\s+equity|^owners.?\s+equity",
                                   bs_col, _sum_um)
                if v: d.total_liabilities = v
    if d.total_equity is None and d.total_assets is not None and d.total_liabilities is not None:
        derived_eq = round(d.total_assets - d.total_liabilities, 2)
        if derived_eq > 0:
            d.total_equity = derived_eq
    # Reverse: derive total_liabilities = total_assets - total_equity.
    # Also use this to fix clearly-wrong extractions (< 0.1% of total assets = from a notes table).
    if d.total_assets is not None and d.total_equity is not None:
        derived_liab = round(d.total_assets - d.total_equity, 2)
        if derived_liab > 0:
            if d.total_liabilities is None:
                d.total_liabilities = derived_liab
            elif d.total_liabilities < d.total_assets * 0.001:
                # total_liabilities < 0.1% of total_assets: almost certainly from a supplemental
                # notes table (e.g. RACE: financial instruments fair value table), not the main BS.
                d.total_liabilities = derived_liab

    ts = extract_shares_from_text(text)
    if not d.shares.common: d.shares.common = ts.common
    if not d.shares.weighted_avg: d.shares.weighted_avg = ts.weighted_avg
    # 从 BS equity 注释提取真实流通股（如 "601 shares issued; 413 and 441 shares outstanding"）
    # 注意：BS equity 里的股本数是原始股数（非货币单位），不应乘以货币 unit_multiplier
    if classified["BS"]:
        for bs_tbl in classified["BS"]:
            for row in bs_tbl["rows"]:
                full_text = " ".join(c for c in row if c)
                # "NNN and NNN shares outstanding"（两列日期时取最新列）
                m_out = re.search(r"([\d,]+)\s*and\s*([\d,]+)\s*shares\s+outstanding", full_text, re.I)
                if m_out:
                    bs_col = detect_latest_column(bs_tbl["rows"])
                    v1 = parse_num(m_out.group(1))
                    v2 = parse_num(m_out.group(2))
                    val = v1 if bs_col == 0 else v2
                    if val and val > 1:
                        # BS equity 中的股本数为原始股数，用 raw_shares_to_m 转换
                        d.shares.common = raw_shares_to_m(val)
                    break
                # "NNN shares outstanding"（排除 authorized/preferred 行）
                m_single = re.search(r"([\d,]+)\s+shares\s+outstanding", full_text, re.I)
                if m_single and not re.search(r"authorized|preferred", full_text, re.I):
                    val = parse_num(m_single.group(1))
                    if val and val > 1:
                        d.shares.common = raw_shares_to_m(val)
                    break
            if d.shares.common: break

    if not d.shares.ads_ratio: d.shares.ads_ratio = ts.ads_ratio
    if not d.shares.ads_per_share and ts.ads_per_share: d.shares.ads_per_share = ts.ads_per_share
    d.shares.details.extend(ts.details)

    # ── 硬编码 ADS 比例兜底（部分中概股的 6-K 不声明比例）────────────────────
    # 来源：各公司官方 20-F / ADS 说明书
    _KNOWN_ADS_RATIOS: Dict[str, float] = {
        "PDD":  4.0,   # Pinduoduo: 1 ADS = 4 ordinary
        "CAN":  15.0,  # Canaan: 1 ADS = 15 ordinary
        "VIPS": 0.2,   # Vipshop: 1 ADS = 0.2 ordinary (1 ord = 5 ADS)
        "MOMO": 2.0,   # Hello Group: 1 ADS = 2 ordinary (changed from 3)
        "IH":   3.0,   # iHuman: 1 ADS = 3 ordinary
        "ATAT": 3.0,   # Atour: 1 ADS = 3 ordinary
        "QFIN": 2.0,   # 360 DigiTech: 1 ADS = 2 ordinary
        "LX":   2.0,   # LexinFintech: 1 ADS = 2 ordinary (changed from 5)
        "TAL":  0.3333, # TAL Education: 3 ADS = 1 ordinary → 1 ADS = 1/3 ordinary
        "JD":   2.0,    # JD.com: 1 ADS = 2 ordinary
        "WDH":  10.0,   # Waterdrop: 1 ADS = 10 Class A ordinary
        "DQ":   5.0,    # Daqo New Energy: 1 ADS = 5 ordinary
        "DDI":  0.05,   # DoubleDown Interactive: 1 ADS = 0.05 ordinary (20 ADS = 1 ordinary)
        "NTES": 5.0,    # NetEase: 1 ADS = 5 ordinary shares (changed from 25 in 2021)
    }
    if not d.shares.ads_per_share:
        _fallback = _KNOWN_ADS_RATIOS.get(d.ticker.upper())
        if _fallback:
            d.shares.ads_per_share = _fallback
            if not d.shares.ads_ratio:
                r = _fallback
                rstr = str(int(r)) if r == int(r) else str(r)
                d.shares.ads_ratio = f"1 ADS = {rstr} shares (hardcoded)"

    # ── DR 比例修正（中概股）──────────────────────────────────────────────────
    # ratio = 每 1 ADS 代表的普通股数量（可 >1 如 PDD=4，也可 <1 如 VIPS=0.2）
    # ADS数量 = 普通股数量 / ratio；每ADS的EPS = 每普通股EPS × ratio
    # 例外：若 IS 表已以 ADS 为单位报告加权平均（wtdavg_already_ads=True），
    #       WtdAvg/Diluted 不再除以 ratio，EPS 也已经是每 ADS 不再换算
    ratio = d.shares.ads_per_share
    if ratio and ratio != 1.0:
        # common（来自 BS，通常是普通股数）始终换算
        if d.shares.common:
            d.shares.common = round(d.shares.common / ratio, 2)
        # WtdAvg/Diluted 仅在 IS 以普通股单位报告时才换算
        if not d.shares.wtdavg_already_ads:
            for attr in ("weighted_avg", "diluted"):
                sv = getattr(d.shares, attr)
                if sv:
                    setattr(d.shares, attr, round(sv / ratio, 2))
            # EPS 换算为每 ADS（仅当换算后仍在合理范围内才做）
            if d.eps_basic is not None and abs(d.eps_basic) * ratio < 500:
                d.eps_basic = round(d.eps_basic * ratio, 4)
            if d.eps_diluted is not None and abs(d.eps_diluted) * ratio < 500:
                d.eps_diluted = round(d.eps_diluted * ratio, 4)

    # 如果 common 仍然为空，用 weighted_avg 基本股数作为近似值
    # （此时 weighted_avg 已经过 ADS 比例修正，可直接作为 ADS 口径使用）
    if not d.shares.common and d.shares.weighted_avg:
        d.shares.common = d.shares.weighted_avg

    # 股本合理性修正：XBRL 文件有时将原始股数（raw count）放在 thousands 单位表中
    # 症状：shares > 30,000M（300亿）时，/ 1000 后更合理（绝大多数公司 < 300亿股）
    _SHARE_MAX_M = 30_000
    for _attr in ("common", "weighted_avg", "diluted"):
        _sv = getattr(d.shares, _attr)
        if _sv and _sv > _SHARE_MAX_M:
            _fixed = round(_sv / 1000, 2)
            if _fixed <= _SHARE_MAX_M:
                setattr(d.shares, _attr, _fixed)
    # 二次校验：basic/common > diluted × 10 说明单位仍错了 1000 倍（如 DSP）
    _ref = d.shares.diluted or d.shares.weighted_avg
    if _ref:
        for _attr in ("weighted_avg", "common"):
            _sv = getattr(d.shares, _attr)
            if _sv and _sv > _ref * 10:
                setattr(d.shares, _attr, round(_sv / 1000, 2))

    # 比率
    if d.revenue and d.revenue != 0:
        if d.gross_profit is not None: d.gross_margin = round(d.gross_profit / d.revenue * 100, 2)
        if d.operating_income is not None: d.operating_margin = round(d.operating_income / d.revenue * 100, 2)
        if d.net_income is not None: d.net_margin = round(d.net_income / d.revenue * 100, 2)
    if d.current_assets and d.current_liabilities and d.current_liabilities != 0:
        d.current_ratio = round(d.current_assets / d.current_liabilities, 2)
    if d.total_equity and d.total_equity != 0:
        # 优先用金融债务（短期+长期借款）；均无时退回总负债
        fin_debt = (d.short_term_debt or 0) + (d.long_term_debt or 0)
        if fin_debt > 0:
            d.debt_to_equity = round(fin_debt / d.total_equity, 2)
        elif d.total_liabilities is not None:
            d.debt_to_equity = round(d.total_liabilities / d.total_equity, 2)
    if d.net_income and d.total_equity and d.total_equity != 0:
        d.roe = round(d.net_income / d.total_equity * 100, 2)
    if d.net_income and d.total_assets and d.total_assets != 0:
        d.roa = round(d.net_income / d.total_assets * 100, 2)
    if d.revenue and d.total_assets and d.total_assets != 0:
        d.asset_turnover = round(d.revenue / d.total_assets, 2)
    if d.ebitda and d.revenue and d.revenue != 0:
        d.ebitda_margin = round(d.ebitda / d.revenue * 100, 2)

    fields = [d.revenue, d.gross_profit, d.operating_income, d.net_income,
              d.total_assets, d.total_equity, d.operating_cf, d.shares.weighted_avg or d.shares.common]
    found = sum(1 for f in fields if f is not None)
    if not _silent:
        print(f"OK {d.filing_type} | {d.currency} {d.unit_label} | {found}/8")
    return d

# ═══════════ 输出 ═══════════
CUR_SYM = {"USD":"$","RMB":"Y","HKD":"HK$","EUR":"E","GBP":"P","JPY":"Y","BRL":"R$"}

def fmt(v, c="USD"):
    if v is None: return "---"
    s, av, sign = CUR_SYM.get(c, c+" "), abs(v), "-" if v < 0 else ""
    if av >= 1e6: return f"{sign}{s}{av/1e6:.1f}T"
    if av >= 1000: return f"{sign}{s}{av/1000:.1f}B"
    if av >= 1: return f"{sign}{s}{av:.1f}M"
    if av >= 0.001: return f"{sign}{s}{av*1000:.0f}K"
    return "---"

def print_summary(results):
    fs = lambda v: "---" if v is None else (f"{v/1000:.2f}B" if abs(v)>=1000 else f"{v:.1f}M")
    pct = lambda v: f"{v:.1f}%" if v is not None else "---"
    print(f"\n{'='*95}")
    print(f"  SEC Filing Analyzer v6.2 (BeautifulSoup)  |  {len(results)} files")
    print(f"{'='*95}")
    for d in results:
        c = d.currency
        print(f"\n+-- {d.ticker} -- {d.filing_type} -- {d.company_name}")
        print(f"|  Date: {d.filing_date}  |  Period: {d.fiscal_period}  |  {c}, {d.unit_label}")
        print(f"|  Revenue: {fmt(d.revenue,c):>14s}  COGS: {fmt(d.cost_of_revenue,c):>14s}  Gross: {fmt(d.gross_profit,c):>14s}  OpInc: {fmt(d.operating_income,c):>14s}")
        print(f"|  R&D: {fmt(d.rd_expense,c):>17s}  S&M: {fmt(d.sm_expense,c):>15s}  G&A: {fmt(d.ga_expense,c):>15s}  SGA: {fmt(d.sga_expense,c):>15s}")
        print(f"|  PreTax: {fmt(d.pretax_income,c):>14s}  Tax: {fmt(d.income_tax,c):>14s}  Net: {fmt(d.net_income,c):>14s}  EBITDA: {fmt(d.ebitda,c):>13s}")
        eb = d.eps_basic if d.eps_basic is not None else "---"
        ed = d.eps_diluted if d.eps_diluted is not None else "---"
        print(f"|  EPS: {eb} / {ed}  IntInc: {fmt(d.interest_income,c)}")
        print(f"|  Assets: {fmt(d.total_assets,c):>14s}  CurAst: {fmt(d.current_assets,c):>13s}  Liab: {fmt(d.total_liabilities,c):>13s}  CurLiab: {fmt(d.current_liabilities,c):>12s}")
        print(f"|  Equity: {fmt(d.total_equity,c):>14s}  Cash: {fmt(d.cash,c):>15s}  STInv: {fmt(d.short_term_investments,c):>13s}  Goodwill: {fmt(d.goodwill,c):>11s}")
        print(f"|  AR: {fmt(d.accounts_receivable,c):>18s}  AP: {fmt(d.accounts_payable,c):>17s}  DefRev: {fmt(d.deferred_revenue,c):>13s}  RetEarn: {fmt(d.retained_earnings,c):>11s}")
        print(f"|  OpCF: {fmt(d.operating_cf,c):>16s}  InvCF: {fmt(d.investing_cf,c):>13s}  FinCF: {fmt(d.financing_cf,c):>13s}  FCF: {fmt(d.free_cash_flow,c):>13s}")
        print(f"|  D&A: {fmt(d.depreciation_amortization,c):>17s}  SBC: {fmt(d.stock_based_compensation,c):>15s}  CapEx: {fmt(d.capex,c):>14s}  Acq: {fmt(d.acquisitions,c):>14s}")
        # Shares：中概股已换算为 ADS 数量（普通股 ÷ ratio），并标注比例
        sh_note = ""
        if d.shares.ads_per_share:
            r = d.shares.ads_per_share
            rstr = str(int(r)) if r == int(r) else str(r)
            sh_note = f"  [ADS, 1={rstr}ord]"
        elif d.shares.ads_ratio:
            sh_note = f"  ({d.shares.ads_ratio})"
        print(f"|  Shares: {fs(d.shares.common):>14s}  WtdAvg: {fs(d.shares.weighted_avg):>12s}  Diluted: {fs(d.shares.diluted):>12s}{sh_note}")
        cr, de = d.current_ratio or "---", d.debt_to_equity or "---"
        print(f"|  Gross: {pct(d.gross_margin):>7s}  Op: {pct(d.operating_margin):>7s}  Net: {pct(d.net_margin):>7s}  EBITDA: {pct(d.ebitda_margin):>6s}")
        print(f"|  ROE: {pct(d.roe):>8s}  ROA: {pct(d.roa):>8s}  Cur: {str(cr):>5}  D/E: {str(de):>5}  ATO: {str(d.asset_turnover or '---'):>5}")
        print(f"+{'-'*93}")
    print()

def write_excel(results, output_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("openpyxl not installed"); write_csv(results, output_path.with_suffix(".csv")); return
    wb = Workbook(); ws = wb.active; ws.title = "Summary"
    hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", fgColor="1E3A5F")
    alt = PatternFill("solid", fgColor="F3F4F6")
    bdr = Border(bottom=Side(style="thin", color="E5E7EB"))
    cols = [("Ticker",10),("Type",8),("Company",28),("Ccy",6),("Unit",18),("Date",14),("Period",16),
            ("Revenue(M)",14),("COGS(M)",12),("Gross(M)",12),("R&D(M)",11),("S&M(M)",11),("G&A(M)",11),
            ("OpEx(M)",12),("OpInc(M)",12),("PreTax(M)",12),("Tax(M)",10),("NetInc(M)",12),
            ("EPS_B",8),("EPS_D",8),("EBITDA(M)",12),
            ("Assets(M)",14),("CurAst(M)",13),("Cash(M)",12),("STInv(M)",12),("AR(M)",10),
            ("Goodwill(M)",11),("Liab(M)",12),("CurLiab(M)",12),("LTDebt(M)",12),
            ("DefRev(M)",11),("AP(M)",10),("Equity(M)",12),("RetEarn(M)",12),("APIC(M)",11),
            ("OpCF(M)",12),("InvCF(M)",12),("FinCF(M)",12),("CapEx(M)",10),("FCF(M)",12),
            ("D&A(M)",10),("SBC(M)",10),("Acq(M)",10),
            ("Shares(M)",11),("WtdAvg(M)",11),("Diluted(M)",11),
            ("Gross%",8),("Op%",7),("Net%",7),("EBITDA%",8),("ROE%",7),("ROA%",7),
            ("CurRatio",8),("D/E",7),("ATO",6)]
    for ci,(n,w) in enumerate(cols,1):
        cell = ws.cell(row=1, column=ci, value=n); cell.font = hf; cell.fill = hfill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = w
    ws.freeze_panes = "A2"
    for ri, d in enumerate(results, 2):
        row = [d.ticker,d.filing_type,d.company_name,d.currency,d.unit_label,d.filing_date,d.fiscal_period,
               d.revenue,d.cost_of_revenue,d.gross_profit,d.rd_expense,d.sm_expense,d.ga_expense,
               d.operating_expenses,d.operating_income,d.pretax_income,d.income_tax,d.net_income,
               d.eps_basic,d.eps_diluted,d.ebitda,
               d.total_assets,d.current_assets,d.cash,d.short_term_investments,d.accounts_receivable,
               d.goodwill,d.total_liabilities,d.current_liabilities,d.long_term_debt,
               d.deferred_revenue,d.accounts_payable,d.total_equity,d.retained_earnings,d.additional_paid_in_capital,
               d.operating_cf,d.investing_cf,d.financing_cf,d.capex,d.free_cash_flow,
               d.depreciation_amortization,d.stock_based_compensation,d.acquisitions,
               d.shares.common,d.shares.weighted_avg,d.shares.diluted,
               d.gross_margin,d.operating_margin,d.net_margin,d.ebitda_margin,d.roe,d.roa,
               d.current_ratio,d.debt_to_equity,d.asset_turnover]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val); cell.border = bdr
            cell.font = Font(name="Arial", size=10)
            if isinstance(val, float) and ci >= 8: cell.number_format = "#,##0.0"
        if ri % 2 == 0:
            for ci in range(1, len(cols)+1): ws.cell(row=ri, column=ci).fill = alt
    output_path.parent.mkdir(parents=True, exist_ok=True); wb.save(str(output_path))
    print(f"\nExcel saved: {output_path}")

def write_csv(results, output_path):
    h = ["Ticker","Type","Company","Ccy","Unit","Date","Period",
         "Revenue(M)","COGS(M)","Gross(M)","R&D(M)","S&M(M)","G&A(M)",
         "OpEx(M)","OpInc(M)","PreTax(M)","Tax(M)","NetInc(M)","EPS_B","EPS_D","EBITDA(M)",
         "Assets(M)","CurAssets(M)","Cash(M)","STInv(M)","AR(M)","Goodwill(M)",
         "Liab(M)","CurLiab(M)","LTDebt(M)","DefRev(M)","AP(M)",
         "Equity(M)","RetEarn(M)","APIC(M)",
         "OpCF(M)","InvCF(M)","FinCF(M)","CapEx(M)","FCF(M)","D&A(M)","SBC(M)","Acq(M)",
         "Shares(M)","WtdAvg(M)","Diluted(M)",
         "Gross%","Op%","Net%","EBITDA%","ROE%","ROA%","CurRatio","D/E","ATO"]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f); w.writerow(h)
        for d in results:
            w.writerow([d.ticker,d.filing_type,d.company_name,d.currency,d.unit_label,
                d.filing_date,d.fiscal_period,
                d.revenue,d.cost_of_revenue,d.gross_profit,d.rd_expense,d.sm_expense,d.ga_expense,
                d.operating_expenses,d.operating_income,d.pretax_income,d.income_tax,d.net_income,
                d.eps_basic,d.eps_diluted,d.ebitda,
                d.total_assets,d.current_assets,d.cash,d.short_term_investments,d.accounts_receivable,
                d.goodwill,d.total_liabilities,d.current_liabilities,d.long_term_debt,
                d.deferred_revenue,d.accounts_payable,d.total_equity,d.retained_earnings,d.additional_paid_in_capital,
                d.operating_cf,d.investing_cf,d.financing_cf,d.capex,d.free_cash_flow,
                d.depreciation_amortization,d.stock_based_compensation,d.acquisitions,
                d.shares.common,d.shares.weighted_avg,d.shares.diluted,
                d.gross_margin,d.operating_margin,d.net_margin,d.ebitda_margin,d.roe,d.roa,
                d.current_ratio,d.debt_to_equity,d.asset_turnover])
    print(f"\nCSV saved: {output_path}")

def write_json(results, output_path):
    import json
    def to_dict(d):
        return {
            "ticker": d.ticker, "filing_type": d.filing_type,
            "company_name": d.company_name, "currency": d.currency,
            "unit_label": d.unit_label, "filing_date": d.filing_date,
            "fiscal_period": d.fiscal_period,
            "income_statement": {
                "revenue": d.revenue, "cost_of_revenue": d.cost_of_revenue,
                "gross_profit": d.gross_profit, "rd_expense": d.rd_expense,
                "sm_expense": d.sm_expense, "ga_expense": d.ga_expense,
                "sga_expense": d.sga_expense, "operating_expenses": d.operating_expenses,
                "operating_income": d.operating_income, "interest_income": d.interest_income,
                "interest_expense": d.interest_expense, "pretax_income": d.pretax_income,
                "income_tax": d.income_tax, "net_income": d.net_income,
                "ebitda": d.ebitda, "eps_basic": d.eps_basic, "eps_diluted": d.eps_diluted,
            },
            "balance_sheet": {
                "total_assets": d.total_assets, "current_assets": d.current_assets,
                "cash": d.cash, "short_term_investments": d.short_term_investments,
                "accounts_receivable": d.accounts_receivable, "inventory": d.inventory,
                "goodwill": d.goodwill, "total_non_current_assets": d.total_non_current_assets,
                "total_liabilities": d.total_liabilities, "current_liabilities": d.current_liabilities,
                "accounts_payable": d.accounts_payable, "deferred_revenue": d.deferred_revenue,
                "long_term_debt": d.long_term_debt,
                "total_non_current_liabilities": d.total_non_current_liabilities,
                "total_equity": d.total_equity, "retained_earnings": d.retained_earnings,
                "additional_paid_in_capital": d.additional_paid_in_capital,
            },
            "cash_flow": {
                "operating_cf": d.operating_cf, "investing_cf": d.investing_cf,
                "financing_cf": d.financing_cf, "capex": d.capex,
                "free_cash_flow": d.free_cash_flow,
                "depreciation_amortization": d.depreciation_amortization,
                "stock_based_compensation": d.stock_based_compensation,
                "change_in_working_capital": d.change_in_working_capital,
                "net_change_in_cash": d.net_change_in_cash,
                "cash_end_of_period": d.cash_end_of_period,
                "acquisitions": d.acquisitions,
            },
            "shares": {
                "common_M": d.shares.common, "weighted_avg_M": d.shares.weighted_avg,
                "diluted_M": d.shares.diluted, "ads_per_share": d.shares.ads_per_share,
            },
            "ratios": {
                "gross_margin_pct": d.gross_margin, "operating_margin_pct": d.operating_margin,
                "net_margin_pct": d.net_margin, "ebitda_margin_pct": d.ebitda_margin,
                "roe_pct": d.roe, "roa_pct": d.roa,
                "current_ratio": d.current_ratio, "debt_to_equity": d.debt_to_equity,
                "asset_turnover": d.asset_turnover,
            },
        }
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump([to_dict(d) for d in results], f, ensure_ascii=False, indent=2)
    print(f"\nJSON saved: {output_path}")

# ═══════════ 并发 worker（须定义在顶层才能被 ProcessPoolExecutor pickle）═══════════
def _worker(args):
    fp, tk = args
    return analyze_filing(fp, tk, _silent=True)

# ═══════════ CLI ═══════════
def main():
    ap = argparse.ArgumentParser(description="SEC Filing Analyzer v6.2 (BeautifulSoup)")
    ap.add_argument("--dir", default=DEFAULT_DIR, help="SEC data directory")
    ap.add_argument("--ticker", nargs="+", help="Filter by ticker(s)")
    ap.add_argument("--file", nargs="+", help="Analyze specific file(s)")
    ap.add_argument("--output", default=None, help="Output file path")
    ap.add_argument("--csv", action="store_true", help="Output CSV instead of Excel")
    ap.add_argument("--workers", type=int, default=1,
                    help="Parallel worker processes (default: 1 = serial)")
    args = ap.parse_args()

    files = []
    if args.file:
        for f in args.file:
            p = Path(f)
            if p.exists(): files.append((p, p.stem.split("_")[0] if "_" in p.stem else ""))
            else: print(f"File not found: {f}")
    else:
        base = Path(args.dir)
        if not base.exists(): print(f"Directory not found: {base.resolve()}"); sys.exit(1)
        for td in sorted(base.iterdir()):
            if not td.is_dir(): continue
            tk = td.name.upper()
            if args.ticker and tk not in [t.upper() for t in args.ticker]: continue
            for fp in sorted(td.glob("*")):
                if fp.suffix.lower() in (".txt",".htm",".html",".xml",".sgml"):
                    if not fp.name.startswith(".") and fp.name != "progress.json":
                        files.append((fp, tk))
        for fp in sorted(base.glob("*")):
            if fp.is_file() and fp.suffix.lower() in (".txt",".htm",".html"):
                if not fp.name.startswith(".") and fp.name != "progress.json":
                    tk = fp.stem.split("_")[0] if "_" in fp.stem else ""
                    if args.ticker and tk.upper() not in [t.upper() for t in args.ticker]: continue
                    files.append((fp, tk))

    if not files: print("No files found"); sys.exit(1)
    workers = max(1, args.workers)
    print(f"\n{'='*60}")
    print(f"  SEC Filing Analyzer v6.2  |  {len(files)} files"
          + (f"  |  {workers} workers" if workers > 1 else ""))
    print(f"{'='*60}\n")
    results = []
    if workers == 1:
        for fp, tk in files:
            try:
                d = analyze_filing(fp, tk)
                if d.filing_type != "Unknown": results.append(d)
            except Exception as e:
                print(f"  ERROR {fp.name}: {e}")
                import traceback; traceback.print_exc()
    else:
        from concurrent.futures import ProcessPoolExecutor, as_completed
        # 用 dict 保持原始顺序
        order = {(fp, tk): i for i, (fp, tk) in enumerate(files)}
        pending = {}
        with ProcessPoolExecutor(max_workers=workers) as exc:
            for fp, tk in files:
                pending[exc.submit(_worker, (fp, tk))] = (fp, tk)
            done = 0
            for fut in as_completed(pending):
                done += 1
                fp, tk = pending[fut]
                try:
                    d = fut.result()
                    status = f"OK {d.filing_type} | {d.currency} {d.unit_label}"
                    if d.filing_type != "Unknown": results.append(d)
                except Exception as e:
                    status = f"ERROR: {e}"
                print(f"  [{done:>{len(str(len(files)))}}/{len(files)}] {tk:8s} {fp.name}  {status}")
        # 按原始文件顺序排列
        results.sort(key=lambda d: next(
            (order[(fp, tk)] for fp, tk in files
             if fp.name == d.file_name and tk == d.ticker), 0))
    if not results: print("No files parsed successfully"); sys.exit(1)
    print_summary(results)
    out = Path(args.output) if args.output else Path(args.dir) / OUTPUT_XLSX
    if args.csv:
        write_csv(results, out.with_suffix(".csv"))
    else:
        write_excel(results, out)
    write_json(results, out.with_suffix(".json"))
    print(f"Done: {len(results)} files\n")

if __name__ == "__main__":
    main()
